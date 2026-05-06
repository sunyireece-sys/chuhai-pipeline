"""
Step5: sales profile enrichment.

Reads step4 `04_verified.xlsx`, fetches richer website context for each eligible
row, extracts contact signals, and asks the LLM to produce a source-grounded
sales profile plus outreach drafts.

Entry rule is intentionally simple:
- Website is present.
- Rating is not "B" (current step4 information-incomplete bucket).

This step does not decide whether a company is a good customer. It enriches the
step4 output so sales and reviewers can judge the profile with evidence.
"""
from __future__ import annotations

import argparse
import asyncio
import datetime as dt
import html
import json
import logging
import os
import re
import time
from dataclasses import asdict, dataclass, field
from pathlib import Path
from urllib.parse import unquote, urljoin, urlparse

from dotenv import load_dotenv
from openpyxl import load_workbook

from schema import write_sales_leads_xlsx

log = logging.getLogger(__name__)

USER_AGENT = (
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
    "AppleWebKit/537.36 (KHTML, like Gecko) "
    "Chrome/146.0.0.0 Safari/537.36"
)

START_PATHS = [
    "",
    "/about",
    "/about-us",
    "/products",
    "/product",
    "/shop",
    "/contact",
    "/contact-us",
    "/wholesale",
    "/bulk",
    "/private-label",
    "/ingredients",
    "/catalog",
]

LINK_KEYWORDS = (
    "about",
    "product",
    "shop",
    "store",
    "catalog",
    "ingredient",
    "wholesale",
    "bulk",
    "private",
    "label",
    "oem",
    "contact",
    "distributor",
    "supply",
)

MAX_PAGE_CHARS = 12000
MAX_LLM_CONTEXT_CHARS = 30000
MAX_DISCOVERED_LINKS = 12
MAX_CRAWL_PAGES = 50
MAX_CRAWL_TEXT_CHARS = 200_000
MAX_SITEMAP_URLS = 10
MIN_PROFILE_EVIDENCE = 3
PROFILE_VERSION = "profile_enrich_v2"
SUPPLIER_PROFILE_PATH = Path(__file__).resolve().parent / "supplier_profile.json"
INTEL_ONLY_NOTICE = "情报留档，不建议主动触达。"
OUTREACH_DRAFT_KEYS = (
    "cold_email_subject",
    "cold_email_body",
    "whatsapp_or_linkedin_message",
    "follow_up_email",
)
CJK_RE = re.compile(r"[\u3400-\u9fff\uF900-\uFAFF]")
SENDER_NAME = "Nicky"
SENDER_TITLE = "Sourcing Manager"
SENDER_COMPANY = "Bairuiyuan Goji"
SENDER_WEBSITE = "http://berylgoji.com"
FIXED_SENDER_SIGNATURE = "\n".join(
    [
        "Best regards,",
        SENDER_NAME,
        f"{SENDER_TITLE}, {SENDER_COMPANY}",
        SENDER_WEBSITE,
    ]
)
FIXED_SENDER_PROFILE = {
    "name": SENDER_NAME,
    "title": SENDER_TITLE,
    "company": SENDER_COMPANY,
    "email_signature": FIXED_SENDER_SIGNATURE,
}
BANNED_WEB_MONITORING_PHRASES = (
    "I " "noticed",
    "I " "noticed this " "detail on your " "website",
    "I " "saw on your " "website",
    "we reviewed your " "website",
    "according to your " "website",
    "came across your company " "website",
    "found your " "website",
)
BANNED_WEB_MONITORING_PHRASES_TEXT = "\n".join(
    f'  - "{phrase}"' for phrase in BANNED_WEB_MONITORING_PHRASES
)
WEB_REFERENCE_OUTREACH_RE = re.compile(
    r"\bI\s+noticed\b|"
    r"\bI\s+saw\b|"
    r"\bwe\s+reviewed\b|"
    r"\byour\s+website\b|"
    r"\bcompany\s+website\b|"
    r"\baccording\s+to\b.{0,80}\bwebsite\b|"
    r"\bthis\s+detail\b|"
    r"\bsource[_\s-]*quote\b|"
    r"\bcrawled\b|"
    r"\bscraped\b",
    re.IGNORECASE | re.DOTALL,
)
REQUIRED_SALES_EMAIL_FACTS = ("Bairuiyuan", "Ningxia", "2003")
REQUIRED_SALES_CTA_HINTS = (
    "target volume",
    "spec",
    "15-minute video call",
    "sourcing or product contact",
    "procurement",
    "sourcing contact",
)
TRACK_PARTNER_LABELS = {
    "skincare_cosmetics": "beauty, skincare, and botanical formulation",
    "nutraceutical_supplements": "nutrition, wellness, and supplement formulation",
    "food_beverage": "functional food, beverage, and clean-label snack",
    "tea_herbal": "tea, herbal, and traditional botanical product",
    "pet_food": "premium pet nutrition and functional treat",
    "other": "nutrition, wellness, functional food, and botanical formulation",
}
TRACK_FALLBACK = "nutraceutical_supplements"
TRACK_LABELS = {
    "skincare_cosmetics": "skincare",
    "nutraceutical_supplements": "supplement formulation",
    "food_beverage": "food and beverage",
    "tea_herbal": "tea and herbal",
    "pet_food": "pet nutrition",
    "other": "supplement formulation",
}
TRACK_KEYWORDS = {
    "skincare_cosmetics": (
        "skincare",
        "skin care",
        "cosmetic",
        "beauty",
        "collagen",
        "serum",
        "cream",
        "skin-brightening",
        "anti-aging",
        "护肤",
        "美妆",
        "化妆",
        "美容",
        "胶原",
    ),
    "nutraceutical_supplements": (
        "supplement",
        "nutraceutical",
        "capsule",
        "tablet",
        "vitamin",
        "immune",
        "eye health",
        "extract",
        "formulation",
        "营养",
        "补充剂",
        "保健",
        "胶囊",
        "片剂",
        "护眼",
        "提取物",
    ),
    "food_beverage": (
        "food",
        "beverage",
        "drink",
        "juice",
        "snack",
        "bakery",
        "cereal",
        "ingredient",
        "superfood",
        "食品",
        "饮料",
        "果汁",
        "零食",
        "烘焙",
        "超级食品",
    ),
    "tea_herbal": (
        "tea",
        "herbal",
        "botanical",
        "infusion",
        "traditional",
        "tcm",
        "herb",
        "茶",
        "草本",
        "花草",
        "中草药",
        "传统",
    ),
    "pet_food": (
        "pet",
        "dog",
        "cat",
        "animal",
        "veterinary",
        "宠物",
        "犬",
        "猫",
        "动物",
    ),
}

FEEDBACK_SCHEMA_FIELDS = [
    "profile_id",
    "profile_version",
    "source_run",
    "sales_user",
    "action_taken",
    "sent_at",
    "reply_status",
    "lead_quality",
    "issue_type",
    "qualitative_feedback",
    "ts",
]

EMAIL_RE = re.compile(
    r"\b[A-Z0-9._%+-]+@[A-Z0-9.-]+\.[A-Z]{2,}\b",
    re.IGNORECASE,
)
PHONE_RE = re.compile(
    r"(?:\+|00)?\d[\d\s()./-]{7,}\d",
)
PHONE_CONTEXT_RE = re.compile(
    r"\b(tel|telephone|phone|hotline|call|mobile|fax|whatsapp|kontakt|telefon|fon)\b",
    re.IGNORECASE,
)
WHATSAPP_RE = re.compile(
    r"https?://(?:wa\.me|api\.whatsapp\.com|chat\.whatsapp\.com)/[^\s\"'<>]+",
    re.IGNORECASE,
)
SOCIAL_RE = re.compile(
    r"https?://(?:www\.)?(?:linkedin\.com|facebook\.com|instagram\.com|x\.com|twitter\.com)/[^\s\"'<>]+",
    re.IGNORECASE,
)
SKIP_URL_EXTENSIONS = (
    ".css",
    ".js",
    ".mjs",
    ".map",
    ".png",
    ".jpg",
    ".jpeg",
    ".gif",
    ".svg",
    ".webp",
    ".ico",
    ".pdf",
    ".doc",
    ".docx",
    ".xls",
    ".xlsx",
    ".zip",
    ".rar",
    ".xml",
    ".json",
    ".webmanifest",
)


@dataclass
class VerifiedLead:
    row_number: int
    input_company_name: str
    input_country: str
    lead_type: str
    xiaoman_company_name: str
    domain: str
    website_country: str
    website: str
    b2b_or_b2c: str
    verified_target: str
    customer_type: str
    is_competitor: str
    goji_presence: str
    rating: str
    p_priority: str
    track_match: str
    matched_track: str
    evidence_url: str
    rating_reason: str
    outreach_angle: str
    contact_count: str
    email_count: str
    contact_name: str
    email: str
    position: str


@dataclass
class PageSnapshot:
    url: str
    status_code: int
    title: str
    text: str
    tel_links: list[str] = field(default_factory=list)
    crawler: str = "httpx"


@dataclass
class ContactSignals:
    emails: list[str]
    phones: list[str]
    phone_candidates_low_confidence: list[str]
    whatsapp_links: list[str]
    contact_pages: list[str]
    social_links: list[str]


def _cell_text(value: object) -> str:
    return str(value or "").strip()


def _load_xlsx_rows(path: Path) -> list[dict]:
    workbook = load_workbook(path, read_only=True)
    sheet = workbook.active
    rows = sheet.iter_rows(values_only=True)
    header_row = next(rows, None)
    if not header_row:
        return []
    header = [_cell_text(value) for value in header_row]
    out: list[dict] = []
    for row_number, raw in enumerate(rows, start=2):
        row = {"_row_number": row_number}
        for idx, key in enumerate(header):
            row[key] = raw[idx] if idx < len(raw) and raw[idx] is not None else ""
        if any(_cell_text(value) for key, value in row.items() if key != "_row_number"):
            out.append(row)
    return out


def load_supplier_profile(path: Path = SUPPLIER_PROFILE_PATH) -> dict:
    if not path.is_file():
        raise FileNotFoundError(
            f"missing supplier profile: {path}. Create supplier_profile.json in the project root before running Step5."
        )
    try:
        profile = json.loads(path.read_text(encoding="utf-8"))
    except json.JSONDecodeError as exc:
        raise ValueError(f"invalid supplier profile JSON: {path}: {exc}") from exc
    if not isinstance(profile, dict):
        raise ValueError(f"supplier profile must be a JSON object: {path}")

    required_fields = (
        "brand_name_en",
        "origin",
        "founded_year",
        "website",
        "positioning",
        "highlights",
        "certifications",
        "product_range",
        "use_case_angles",
    )
    missing = [field_name for field_name in required_fields if not profile.get(field_name)]
    if missing:
        raise ValueError(f"supplier profile missing required fields: {', '.join(missing)}")
    use_case_angles = profile.get("use_case_angles")
    if not isinstance(use_case_angles, dict) or TRACK_FALLBACK not in use_case_angles:
        raise ValueError(
            "supplier profile use_case_angles must include nutraceutical_supplements fallback"
        )
    return profile


def _truth_text(value: object) -> str:
    if isinstance(value, bool):
        return "Yes" if value else "No"
    return _cell_text(value)


def _lead_from_row(row: dict) -> VerifiedLead:
    return VerifiedLead(
        row_number=int(row.get("_row_number") or 0),
        input_company_name=_cell_text(row.get("Input Company Name")),
        input_country=_cell_text(row.get("Input Country")),
        lead_type=_cell_text(row.get("Lead Type")),
        xiaoman_company_name=_cell_text(row.get("Xiaoman Company Name")),
        domain=_cell_text(row.get("Domain")),
        website_country=_cell_text(row.get("Website Country")),
        website=_cell_text(row.get("Website")),
        b2b_or_b2c=_cell_text(row.get("B2B/B2C")),
        verified_target=_cell_text(row.get("Verified Target")),
        customer_type=_cell_text(row.get("Customer Type")),
        is_competitor=_truth_text(row.get("Is Competitor")),
        goji_presence=_cell_text(row.get("Goji Presence")),
        rating=_cell_text(row.get("Rating")),
        p_priority=_cell_text(row.get("P Priority")),
        track_match=_cell_text(row.get("Track Match")),
        matched_track=_cell_text(row.get("Matched Track")),
        evidence_url=_cell_text(row.get("Evidence URL")),
        rating_reason=_cell_text(row.get("Rating Reason")),
        outreach_angle=_cell_text(row.get("Outreach Angle")),
        contact_count=_cell_text(row.get("Contact Count")),
        email_count=_cell_text(row.get("Email Count")),
        contact_name=_cell_text(row.get("Contact Name")),
        email=_cell_text(row.get("Email")),
        position=_cell_text(row.get("Position")),
    )


def _resolved_country_for_lead(lead: VerifiedLead) -> str:
    website_country = _cell_text(lead.website_country)
    if website_country and website_country.lower() != "unclear":
        return website_country
    return lead.input_country


def load_enrich_leads(verified_xlsx: Path) -> list[VerifiedLead]:
    """Return step4 rows that should enter enrichment."""
    leads: list[VerifiedLead] = []
    for row in _load_xlsx_rows(verified_xlsx):
        lead = _lead_from_row(row)
        if not lead.website:
            continue
        if lead.rating.upper() == "B":
            continue
        leads.append(lead)
    return leads


def _base_url(value: str) -> str:
    text = _cell_text(value)
    if not text:
        return ""
    if text.startswith(("http://", "https://")):
        parsed = urlparse(text)
        return f"{parsed.scheme}://{parsed.netloc}"
    return f"https://{text.strip('/')}"


def _crawl_base_from_final_url(value: str) -> str:
    parsed = urlparse(value)
    if not parsed.scheme or not parsed.netloc:
        return _base_url(value)
    path = parsed.path or ""
    if path and not path.endswith("/"):
        path = path.rsplit("/", 1)[0] + "/"
    if path == "/":
        path = ""
    return f"{parsed.scheme}://{parsed.netloc}{path}".rstrip("/")


def _slugify(value: str, fallback: str = "company") -> str:
    text = value.lower()
    text = re.sub(r"https?://", "", text)
    text = re.sub(r"[^a-z0-9]+", "-", text)
    text = text.strip("-")
    return (text or fallback)[:80]


def _lead_slug(lead: VerifiedLead) -> str:
    basis = lead.xiaoman_company_name or lead.input_company_name or lead.website
    return f"{lead.row_number:03d}-{_slugify(basis)}"


def _html_to_text(content: str) -> str:
    content = re.sub(r"(?is)<(script|style|noscript|svg).*?</\1>", " ", content)
    content = re.sub(r"(?is)<br\s*/?>", "\n", content)
    content = re.sub(r"(?is)</(p|div|li|h[1-6]|tr)>", "\n", content)
    content = re.sub(r"(?s)<[^>]+>", " ", content)
    content = html.unescape(content)
    lines = [re.sub(r"\s+", " ", line).strip() for line in content.splitlines()]
    return "\n".join(line for line in lines if line)


def _page_title(content: str) -> str:
    match = re.search(r"(?is)<title[^>]*>(.*?)</title>", content)
    if not match:
        return ""
    return re.sub(r"\s+", " ", html.unescape(match.group(1))).strip()


def _extract_links(content: str, base_url: str) -> list[str]:
    urls: list[str] = []
    seen: set[str] = set()
    base_host = urlparse(base_url).netloc.lower()
    for raw in re.findall(r"""href=["']([^"']+)["']""", content, flags=re.IGNORECASE):
        if raw.startswith(("mailto:", "tel:", "javascript:", "#")):
            continue
        absolute = urljoin(base_url + "/", raw)
        parsed = urlparse(absolute)
        if parsed.scheme not in {"http", "https"}:
            continue
        if parsed.netloc.lower() != base_host:
            continue
        path = parsed.path.lower()
        if path.endswith(SKIP_URL_EXTENSIONS) or "/wp-json/" in path:
            continue
        cleaned = parsed._replace(fragment="", query="").geturl().rstrip("/")
        lowered = cleaned.lower()
        if cleaned in seen:
            continue
        if not any(keyword in lowered for keyword in LINK_KEYWORDS):
            continue
        seen.add(cleaned)
        urls.append(cleaned)
    return urls[:MAX_DISCOVERED_LINKS]


def _extract_tel_links(content: str) -> list[str]:
    values: list[str] = []
    for raw in re.findall(r"""href=["']\s*tel:([^"']+)["']""", content, flags=re.IGNORECASE):
        cleaned = unquote(raw).split("?", 1)[0].strip()
        if cleaned:
            values.append(cleaned)
    return _unique(values)


def _is_htmlish_url(url: str) -> bool:
    path = urlparse(url).path.lower()
    if not path:
        return True
    return not path.endswith(SKIP_URL_EXTENSIONS)


def _same_host_url(url: str, base_host: str) -> str:
    parsed = urlparse(url)
    if parsed.scheme not in {"http", "https"}:
        return ""
    if parsed.netloc.lower() != base_host.lower():
        return ""
    lowered_url = url.lower()
    if "{" in lowered_url or "%7b" in lowered_url:
        return ""
    if not _is_htmlish_url(url):
        return ""
    return parsed._replace(fragment="", query="").geturl().rstrip("/")


def _append_url(urls: list[str], seen: set[str], url: str) -> None:
    if url and url not in seen:
        seen.add(url)
        urls.append(url)


def _discover_sitemap_urls(base_url: str, client: object) -> list[str]:
    base_host = urlparse(base_url).netloc.lower()
    discovered: list[str] = []
    seen: set[str] = set()
    for path in ("/sitemap.xml", "/sitemap_index.xml"):
        sitemap_url = urljoin(base_url + "/", path.lstrip("/"))
        try:
            response = client.get(sitemap_url)
            response.raise_for_status()
        except Exception:
            continue
        for raw in re.findall(r"(?is)<loc>\s*([^<]+)\s*</loc>", response.text):
            cleaned = html.unescape(raw).strip()
            same_host = _same_host_url(cleaned, base_host)
            if same_host:
                _append_url(discovered, seen, same_host)
    priority = [
        url
        for url in discovered
        if any(keyword in url.lower() for keyword in LINK_KEYWORDS)
    ]
    return priority[: min(MAX_SITEMAP_URLS, max(0, MAX_CRAWL_PAGES - 1))]


def _planned_crawl_urls(home_url: str, home_html: str, client: object) -> list[str]:
    crawl_base = _crawl_base_from_final_url(home_url)
    seen: set[str] = {home_url.rstrip("/")}
    planned: list[str] = []
    for path in START_PATHS[1:]:
        _append_url(planned, seen, urljoin(crawl_base + "/", path.lstrip("/")).rstrip("/"))
    for discovered in _extract_links(home_html, crawl_base):
        _append_url(planned, seen, discovered)
    for sitemap_url in _discover_sitemap_urls(crawl_base, client):
        _append_url(planned, seen, sitemap_url)
    return planned[: max(0, MAX_CRAWL_PAGES - 1)]


def _limited_page(page: PageSnapshot, remaining_chars: int) -> PageSnapshot | None:
    if remaining_chars <= 0 or not page.text:
        return None
    page.text = page.text[: min(MAX_PAGE_CHARS, remaining_chars)]
    return page


def _add_page(pages: list[PageSnapshot], page: PageSnapshot | None, total_chars: int) -> int:
    if not page:
        return total_chars
    if any(existing.url == page.url for existing in pages):
        return total_chars
    limited = _limited_page(page, MAX_CRAWL_TEXT_CHARS - total_chars)
    if not limited:
        return total_chars
    pages.append(limited)
    return total_chars + len(limited.text)


def _page_from_html(
    *,
    url: str,
    status_code: int,
    html_content: str,
    crawler: str,
    fallback_text: str = "",
) -> PageSnapshot | None:
    text = _html_to_text(html_content) if html_content else fallback_text.strip()
    if not text:
        return None
    return PageSnapshot(
        url=url.rstrip("/"),
        status_code=status_code,
        title=_page_title(html_content),
        text=text,
        tel_links=_extract_tel_links(html_content),
        crawler=crawler,
    )


def _crawl4ai_result_text(result: object) -> str:
    markdown = getattr(result, "markdown", "") or ""
    if isinstance(markdown, str):
        return markdown
    raw_markdown = getattr(markdown, "raw_markdown", "") or getattr(markdown, "fit_markdown", "")
    if raw_markdown:
        return str(raw_markdown)
    return str(markdown) if markdown else ""


def _crawl4ai_result_html(result: object) -> str:
    return str(getattr(result, "cleaned_html", "") or getattr(result, "html", "") or "")


def _crawl4ai_result_title(result: object, html_content: str) -> str:
    metadata = getattr(result, "metadata", {}) or {}
    if isinstance(metadata, dict):
        title = _cell_text(metadata.get("title"))
        if title:
            return title
    return _page_title(html_content)


def _page_from_crawl4ai_result(result: object, requested_url: str) -> PageSnapshot | None:
    if getattr(result, "success", True) is False:
        return None
    html_content = _crawl4ai_result_html(result)
    text = _html_to_text(html_content) if html_content else _crawl4ai_result_text(result).strip()
    if not text:
        return None
    final_url = str(getattr(result, "url", requested_url) or requested_url).rstrip("/")
    status_code = int(getattr(result, "status_code", 0) or 0)
    return PageSnapshot(
        url=final_url,
        status_code=status_code,
        title=_crawl4ai_result_title(result, html_content),
        text=text,
        tel_links=_extract_tel_links(html_content),
        crawler="crawl4ai",
    )


def _crawl4ai_configs(timeout_s: float) -> tuple[object, object]:
    from crawl4ai import BrowserConfig, CrawlerRunConfig

    browser_config_kwargs = {"headless": True}
    try:
        browser_config = BrowserConfig(**browser_config_kwargs)
    except TypeError:
        browser_config = BrowserConfig()

    run_config_kwargs = {
        "word_count_threshold": 5,
        "exclude_external_links": True,
        "remove_overlay_elements": True,
        "page_timeout": int(timeout_s * 1000),
    }
    try:
        from crawl4ai import CacheMode

        run_config_kwargs["cache_mode"] = CacheMode.BYPASS
    except Exception:
        pass
    try:
        run_config = CrawlerRunConfig(**run_config_kwargs)
    except TypeError:
        run_config = CrawlerRunConfig()
    return browser_config, run_config


async def _crawl4ai_arun(crawler: object, url: str, run_config: object) -> object:
    try:
        return await crawler.arun(url=url, config=run_config)
    except TypeError:
        return await crawler.arun(url=url)


async def _fetch_site_pages_crawl4ai_async(website: str, *, timeout_s: float) -> list[PageSnapshot]:
    import httpx
    from crawl4ai import AsyncWebCrawler

    base = _base_url(website)
    if not base:
        return []

    browser_config, run_config = _crawl4ai_configs(timeout_s)
    try:
        crawler_context = AsyncWebCrawler(config=browser_config)
    except TypeError:
        crawler_context = AsyncWebCrawler()

    pages: list[PageSnapshot] = []
    total_chars = 0
    async with crawler_context as crawler:
        home_result = await _crawl4ai_arun(crawler, base, run_config)
        home_page = _page_from_crawl4ai_result(home_result, base)
        total_chars = _add_page(pages, home_page, total_chars)
        if not pages:
            return []

        home_html = _crawl4ai_result_html(home_result)
        home_url = pages[0].url
        with httpx.Client(
            follow_redirects=True,
            timeout=timeout_s,
            headers={"User-Agent": USER_AGENT},
        ) as client:
            planned_urls = _planned_crawl_urls(home_url, home_html, client)

        for url in planned_urls:
            if len(pages) >= MAX_CRAWL_PAGES or total_chars >= MAX_CRAWL_TEXT_CHARS:
                break
            try:
                result = await _crawl4ai_arun(crawler, url, run_config)
            except Exception as exc:
                log.info("crawl4ai fetch failed %s: %s", url, exc)
                continue
            total_chars = _add_page(pages, _page_from_crawl4ai_result(result, url), total_chars)
    return pages


def _fetch_site_pages_crawl4ai(website: str, *, timeout_s: float) -> list[PageSnapshot]:
    try:
        return asyncio.run(_fetch_site_pages_crawl4ai_async(website, timeout_s=timeout_s))
    except ImportError:
        log.info("crawl4ai unavailable; falling back to httpx crawler")
        return []
    except Exception as exc:
        log.info("crawl4ai failed for %s: %s; falling back to httpx crawler", website, exc)
        return []


def fetch_site_pages_httpx(website: str, *, timeout_s: float = 15.0) -> list[PageSnapshot]:
    import httpx

    base = _base_url(website)
    if not base:
        return []

    pages: list[PageSnapshot] = []
    with httpx.Client(
        follow_redirects=True,
        timeout=timeout_s,
        headers={"User-Agent": USER_AGENT},
    ) as client:
        try:
            home_response = client.get(base)
            home_response.raise_for_status()
        except Exception as exc:
            log.info("fetch failed %s: %s", base, exc)
            return []

        content_type = home_response.headers.get("content-type", "").lower()
        if content_type and "text/html" not in content_type and "application/xhtml" not in content_type:
            log.info("fetch skipped non-html %s: %s", base, content_type)
            return []

        home_text = _html_to_text(home_response.text)
        final_home_url = str(home_response.url).rstrip("/")
        total_chars = 0
        if home_text:
            total_chars = _add_page(
                pages,
                PageSnapshot(
                    url=final_home_url,
                    status_code=home_response.status_code,
                    title=_page_title(home_response.text),
                    text=home_text,
                    tel_links=_extract_tel_links(home_response.text),
                    crawler="httpx",
                ),
                total_chars,
            )

        planned_urls = _planned_crawl_urls(final_home_url, home_response.text, client)

        idx = 0
        while idx < len(planned_urls):
            if len(pages) >= MAX_CRAWL_PAGES or total_chars >= MAX_CRAWL_TEXT_CHARS:
                break
            url = planned_urls[idx]
            idx += 1
            try:
                response = client.get(url)
                response.raise_for_status()
            except Exception as exc:
                log.info("fetch failed %s: %s", url, exc)
                continue
            content_type = response.headers.get("content-type", "").lower()
            if content_type and "text/html" not in content_type and "application/xhtml" not in content_type:
                log.info("fetch skipped non-html %s: %s", url, content_type)
                continue
            text = _html_to_text(response.text)
            if not text:
                continue
            final_url = str(response.url).rstrip("/")
            total_chars = _add_page(
                pages,
                PageSnapshot(
                    url=final_url,
                    status_code=response.status_code,
                    title=_page_title(response.text),
                    text=text,
                    tel_links=_extract_tel_links(response.text),
                    crawler="httpx",
                ),
                total_chars,
            )
    return pages


def fetch_site_pages(website: str, *, timeout_s: float = 15.0) -> list[PageSnapshot]:
    pages = _fetch_site_pages_crawl4ai(website, timeout_s=timeout_s)
    if pages:
        return pages
    return fetch_site_pages_httpx(website, timeout_s=timeout_s)


def _unique(values: list[str]) -> list[str]:
    seen: set[str] = set()
    out: list[str] = []
    for value in values:
        cleaned = value.strip().strip(".,;:()[]{}<>")
        if not cleaned:
            continue
        lowered = cleaned.lower()
        if lowered in seen:
            continue
        seen.add(lowered)
        out.append(cleaned)
    return out


def _unique_phone_values(values: list[str]) -> list[str]:
    seen: set[str] = set()
    out: list[str] = []
    for value in values:
        cleaned = value.strip().strip(".,;:[]{}<>")
        if not cleaned:
            continue
        key = _phone_digits(cleaned) or cleaned.lower()
        if key in seen:
            continue
        seen.add(key)
        out.append(cleaned)
    return out


def _phone_digits(value: str) -> str:
    return re.sub(r"\D+", "", value or "")


def _clean_phone_candidate(value: str) -> str:
    text = re.sub(r"\s+", " ", value or "").strip()
    return text.strip(".,;:[]{}<>")


def _has_strong_phone_format(value: str) -> bool:
    return bool(
        re.search(r"(\+|00|\(\d{2,4}\)|\d{2,}\s*[-/]\s*\d{2,})", value)
    )


def _classify_phone_candidate(raw: str, *, source: str, context: str = "") -> tuple[str, str] | None:
    cleaned = _clean_phone_candidate(raw)
    digits = _phone_digits(cleaned)
    if not cleaned or len(digits) > 16:
        return None
    sku_like_digits = re.sub(r"[\s\-()]+", "", cleaned)
    if (
        len(sku_like_digits) >= 14
        and sku_like_digits.isdigit()
        and not cleaned.lstrip().startswith("+")
        and not re.search(r"[\s-]", cleaned)
    ):
        return cleaned, "low"
    if source == "tel":
        if len(digits) >= 7:
            return cleaned, "high"
        return cleaned, "low"

    context_has_phone_label = bool(PHONE_CONTEXT_RE.search(context or ""))
    strong_format = _has_strong_phone_format(cleaned)
    if len(digits) < 8:
        return cleaned, "low"
    if "\n" in raw and not context_has_phone_label and not strong_format:
        return cleaned, "low"
    if len(digits) >= 10 and (context_has_phone_label or strong_format):
        return cleaned, "high"
    if len(digits) >= 8 and context_has_phone_label and strong_format:
        return cleaned, "high"
    return cleaned, "low"


def _split_phone_confidence(pages: list[PageSnapshot]) -> tuple[list[str], list[str]]:
    high: list[str] = []
    low: list[str] = []
    for page in pages:
        for raw in page.tel_links:
            classified = _classify_phone_candidate(raw, source="tel")
            if not classified:
                continue
            cleaned, confidence = classified
            (high if confidence == "high" else low).append(cleaned)
        for match in PHONE_RE.finditer(page.text):
            context = page.text[max(0, match.start() - 80) : match.end() + 80]
            classified = _classify_phone_candidate(match.group(0), source="text", context=context)
            if not classified:
                continue
            cleaned, confidence = classified
            (high if confidence == "high" else low).append(cleaned)
    high_unique = _unique_phone_values(high)
    high_digits = {_phone_digits(phone) for phone in high_unique}
    low_unique = [
        phone
        for phone in _unique_phone_values(low)
        if _phone_digits(phone) not in high_digits
    ]
    return high_unique, low_unique


def extract_contacts(pages: list[PageSnapshot], lead: VerifiedLead) -> ContactSignals:
    text = "\n".join(page.text for page in pages)
    emails = _unique(EMAIL_RE.findall(text) + [part.strip() for part in lead.email.split(";")])
    phones, phone_candidates_low_confidence = _split_phone_confidence(pages)
    whatsapp_links = _unique(WHATSAPP_RE.findall(text))
    social_links = _unique(SOCIAL_RE.findall(text))
    contact_pages = _unique(
        [
            page.url
            for page in pages
            if "contact" in page.url.lower()
            or "contact" in page.title.lower()
            or "contact" in page.text[:1500].lower()
        ]
    )
    return ContactSignals(
        emails=emails,
        phones=phones,
        phone_candidates_low_confidence=phone_candidates_low_confidence,
        whatsapp_links=whatsapp_links,
        contact_pages=contact_pages,
        social_links=social_links,
    )


def render_site_markdown(lead: VerifiedLead, pages: list[PageSnapshot]) -> str:
    lines = [
        f"# Raw Website Text: {lead.xiaoman_company_name or lead.input_company_name}",
        "",
        f"- Input Company: {lead.input_company_name}",
        f"- Xiaoman Company: {lead.xiaoman_company_name}",
        f"- Website: {lead.website}",
        f"- Generated At: {dt.datetime.now().isoformat(timespec='seconds')}",
        "",
    ]
    for page in pages:
        lines.extend(
            [
                f"## {page.url}",
                "",
                f"- Status: {page.status_code}",
                f"- Title: {page.title}",
                "",
                page.text,
                "",
            ]
        )
    return "\n".join(lines).rstrip() + "\n"


def _llm_context(
    lead: VerifiedLead,
    pages: list[PageSnapshot],
    contacts: ContactSignals,
    supplier_profile: dict,
) -> dict:
    page_blocks = []
    remaining = MAX_LLM_CONTEXT_CHARS
    for page in pages:
        block = f"SOURCE_URL: {page.url}\nTITLE: {page.title}\nTEXT:\n{page.text}\n"
        if remaining <= 0:
            break
        page_blocks.append(block[:remaining])
        remaining -= len(page_blocks[-1])
    return {
        "lead": asdict(lead),
        "contact_signals": asdict(contacts),
        "supplier_profile": supplier_profile,
        "sender_identity": FIXED_SENDER_PROFILE,
        "website_pages": page_blocks,
    }


PROFILE_SYSTEM_PROMPT = """
You create source-grounded B2B profile cards for goji berry export review and outreach.
Return ONLY valid JSON. Do not include markdown.

Rules:
- Use only facts from the provided step4 row, contact signals, website pages,
  supplier_profile, and sender_identity.
- Every target-company business claim recorded in internal profile evidence must include a
  source_url and source_quote.
- If a fact is not supported, write "未找到证据" instead of guessing.
- Do not hide bad facts. If step4 says the company is a competitor or not a target,
  reflect that in internal_notes and outreach_risk.
- If step4 says the company is a competitor or not a target, do not generate sales
  outreach. Set outreach draft fields to "" and explain "情报留档，不建议主动触达".
- Aim for at least 3 evidence items when the sources support them. Cover as many
  of these evidence_type values as possible: identity, business, product_goji,
  contact, outreach_hook. Do not fabricate evidence to hit the count.
- Outreach drafts must not invent target-company certifications, purchase
  intent, decision makers, volumes, pricing, production capacity, or existing
  relationship. Only supplier certifications listed in supplier_profile may be
  mentioned for Bairuiyuan.
- Outreach drafts must use only a generalized business direction inferred from
  the profile, such as nutrition, wellness, functional food, botanical
  formulation, skincare, tea/herbal, or pet nutrition. Do not copy or expose
  source quotes, page titles, raw crawler text, or page-specific observations.
- Use supplier_profile as the supplier source of truth. The outreach product
  source is Bairuiyuan goji from Ningxia, China.
- Use this sender identity exactly, with literal text rather than placeholders:
  Nicky, Sourcing Manager, Bairuiyuan Goji.
- Do not use {{sender.name}}, {{sender.company}}, {{sender.title}},
  {{sender.email_signature}}, [Your Name], or any other sender placeholder.
- Write Chinese analysis fields for internal review.
- The four outreach draft fields must be English only. Do not use Chinese
  characters in the subject, email body, WhatsApp/LinkedIn message, or
  follow-up email. If evidence is in Chinese or another language, paraphrase it
  in English for outreach.
- Before writing outreach, infer the target company's business track from
  profile.bio_cn and profile.business_relevance_cn. Use exactly one of:
  skincare_cosmetics, nutraceutical_supplements, food_beverage, tea_herbal,
  pet_food, other. If the track is other, use nutraceutical_supplements as the
  supplier hook fallback.
- Choose the matching supplier_profile.use_case_angles value and make it the
  Bairuiyuan capability hook in cold_email_body.
- After the greeting, open naturally as a foreign-trade sourcing email. Ask
  whether the recipient's team is currently sourcing goji berries or goji
  extract for related product lines. It is acceptable to mention the generalized
  business track, but do not mention exact page details or quote source text.
- Outreach drafts must never create a monitoring, scraping, audit, or direct
  quotation feeling. Banned phrasings include:
{BANNED_WEB_MONITORING_PHRASES_TEXT}
- cold_email_body must introduce Bairuiyuan as founded in 2003 in Ningxia, with
  self-managed organic goji plantations and goji extract capability. Mention
  relevant certifications by geography: for Europe emphasize EU BCS Organic;
  for the United States or North America emphasize FDA and KOSHER; for the
  Middle East emphasize HALAL.
- cold_email_body must invite the recipient to reply with volume/spec needs or
  schedule a 15-minute video call.
- cold_email_body and follow_up_email must end with this exact signature:
  Best regards,
  Nicky
  Sourcing Manager, Bairuiyuan Goji
  http://berylgoji.com
- cold_email_subject must include the target company name and a business-track
  keyword. Examples: "Ningxia organic goji for [company] skincare line" or
  "Bairuiyuan goji extract — supplement formulation partner for [company]".
- whatsapp_or_linkedin_message must be 80 words or fewer, ask about sourcing or
  procurement needs, and avoid page-specific observations.
- follow_up_email must be 120 words or fewer, reference the first email subject,
  and introduce one new angle such as a spec sheet, sample, or video call.

Output JSON schema:
{
  "company": {
    "display_name": string,
    "input_company_name": string,
    "website": string,
    "country": string,
    "step4_rating": string,
    "step4_customer_type": string,
    "step4_goji_presence": string
  },
  "profile": {
    "bio_cn": string,
    "bio_en": string,
    "business_relevance_cn": string,
    "goji_or_adjacent_evidence": [
      {
        "evidence_type": "identity" | "business" | "product_goji" | "contact" | "outreach_hook",
        "claim": string,
        "source_url": string,
        "page_section": string,
        "source_quote": string
      }
    ],
    "contact_summary_cn": string,
    "recommended_outreach_angle_cn": string,
    "missing_info_cn": [string],
    "outreach_risk_cn": [string],
    "internal_notes_cn": string
  },
  "outreach": {
    "cold_email_subject": string,
    "cold_email_body": string,
    "whatsapp_or_linkedin_message": string,
    "follow_up_email": string
  }
}
""".strip().replace("{BANNED_WEB_MONITORING_PHRASES_TEXT}", BANNED_WEB_MONITORING_PHRASES_TEXT)

NICKY_STYLE_SYSTEM_PROMPT = """
You are a senior B2B foreign-trade copy editor for an exporter of Ningxia goji
ingredients. You receive four English outreach drafts and rewrite them in
polished, Nicky-style sourcing English. Preserve company names, supplier facts,
certification claims, the exact sender signature, all length caps, and the
generalized business direction. Do not preserve page-specific quoted material
or web-monitoring phrasing. Only change phrasing and tone.

Style targets:
- Natural foreign-trade sourcing tone: state the purpose, ask whether goji
  berries or goji extract are relevant to current sourcing needs, and introduce
  Bairuiyuan's capability without sounding like a web audit.
- "This is {NAME} of {COMPANY}" or "My name is {NAME} with {COMPANY}"
  instead of "I am {NAME} from {COMPANY}".
- "I am reaching out to ask whether your team is currently sourcing goji berries
  or goji extract" or "I wanted to ask whether this category is relevant to
  your current sourcing work" for the relevance ask.
- "Could you connect me with the sourcing or product contact responsible for
  botanical ingredients?" for the CTA, when appropriate.
- WhatsApp / LinkedIn message must remain short, conversational, single
  paragraph, under 80 words.
- Follow-up email must remain under 120 words.

Banned phrasings (rewrite if present):
- Any language that says or implies the sender inspected pages, scraped content,
  reviewed page details, or is quoting source material.
{BANNED_WEB_MONITORING_PHRASES_TEXT}
- "we are writing to enquire whether your organization has sourcing demands"
- "potential goji berry buyers"
- "would like to understand whether this category is relevant"
- "could you direct me to the person"

Hard constraints:
- Do NOT translate to or insert any Chinese characters.
- Do NOT directly quote source text, page titles, crawler text, or evidence
  source_quote content in the outreach drafts.
- Remove all web-monitoring, crawler-like, and direct-quotation phrasing.
- Do NOT remove or alter the signature block:
  Best regards,
  Nicky
  Sourcing Manager, Bairuiyuan Goji
  http://berylgoji.com
- Do NOT change company names, certification names, founding year (2003),
  Ningxia, or any factual claim.
- Do NOT add new factual claims that were not in the input.

Return ONLY valid JSON with these four string keys, no markdown:
cold_email_subject, cold_email_body, whatsapp_or_linkedin_message, follow_up_email.
""".strip().replace("{BANNED_WEB_MONITORING_PHRASES_TEXT}", BANNED_WEB_MONITORING_PHRASES_TEXT)


def _json_loads_loose(text: str) -> dict:
    text = text.strip()
    if text.startswith("```"):
        text = re.sub(r"^```(?:json)?\s*", "", text)
        text = re.sub(r"\s*```$", "", text)
    return json.loads(text)


def synthesize_profile(
    lead: VerifiedLead,
    pages: list[PageSnapshot],
    contacts: ContactSignals,
    supplier_profile: dict,
) -> dict:
    api_key = (
        os.environ.get("LLM_API_KEY")
        or os.environ.get("OPENAI_API_KEY")
        or os.environ.get("GLM_API_KEY")
    )
    if not api_key:
        raise RuntimeError("LLM_API_KEY is empty. Set it in .env before running profile synthesis.")

    from openai import OpenAI

    client_kwargs: dict = {"api_key": api_key}
    base_url = os.environ.get("LLM_BASE_URL")
    if base_url:
        client_kwargs["base_url"] = base_url
    client_kwargs["timeout"] = float(os.environ.get("LLM_TIMEOUT_SECONDS", "75"))
    client = OpenAI(**client_kwargs)
    response = client.chat.completions.create(
        model=os.environ.get("LLM_MODEL", "gpt-4o-mini"),
        response_format={"type": "json_object"},
        messages=[
            {"role": "system", "content": PROFILE_SYSTEM_PROMPT},
            {
                "role": "user",
                "content": json.dumps(
                    _llm_context(lead, pages, contacts, supplier_profile),
                    ensure_ascii=False,
                ),
            },
        ],
        max_tokens=int(os.environ.get("PROFILE_LLM_MAX_TOKENS", "2500")),
        temperature=0,
    )
    return _json_loads_loose(response.choices[0].message.content or "{}")


def rewrite_outreach_tone(outreach: dict) -> dict:
    """Rewrite sales outreach in Nicky's tone; keep pass-1 output on failure."""
    keys = (
        "cold_email_subject",
        "cold_email_body",
        "whatsapp_or_linkedin_message",
        "follow_up_email",
    )
    payload = {key: _cell_text(outreach.get(key)) for key in keys}
    if not any(payload.values()):
        return outreach

    api_key = (
        os.environ.get("LLM_API_KEY")
        or os.environ.get("OPENAI_API_KEY")
        or os.environ.get("GLM_API_KEY")
    )
    if not api_key:
        return outreach

    try:
        from openai import OpenAI

        client_kwargs: dict = {"api_key": api_key}
        base_url = os.environ.get("LLM_BASE_URL")
        if base_url:
            client_kwargs["base_url"] = base_url
        client_kwargs["timeout"] = float(os.environ.get("LLM_TIMEOUT_SECONDS", "75"))
        client = OpenAI(**client_kwargs)
        response = client.chat.completions.create(
            model=os.environ.get("LLM_MODEL", "gpt-4o-mini"),
            response_format={"type": "json_object"},
            messages=[
                {"role": "system", "content": NICKY_STYLE_SYSTEM_PROMPT},
                {"role": "user", "content": json.dumps(payload, ensure_ascii=False)},
            ],
            max_tokens=int(os.environ.get("PROFILE_LLM_MAX_TOKENS", "2500")),
            temperature=0.2,
        )
        rewritten = _json_loads_loose(response.choices[0].message.content or "{}")
    except Exception as exc:
        log.warning("Nicky tone rewrite failed, keeping pass-1 outreach: %s", exc)
        return outreach

    out = dict(outreach)
    for key in keys:
        new_value = _cell_text(rewritten.get(key))
        if not new_value:
            continue
        if _contains_cjk(new_value):
            log.warning("Nicky rewrite produced CJK in %s, falling back for that field", key)
            continue
        out[key] = new_value
    return out


def _is_yes(value: str) -> bool:
    return _cell_text(value).lower() in {"yes", "true", "1", "y", "是"}


def _is_no(value: str) -> bool:
    return _cell_text(value).lower() in {"no", "false", "0", "n", "否"}


def _outreach_mode_for_lead(lead: VerifiedLead) -> str:
    if _is_yes(lead.is_competitor) or lead.customer_type == "竞争对手" or lead.rating.upper() == "Z":
        return "intel_only"
    if _is_no(lead.verified_target):
        return "intel_only"
    return "sales"


def _blank_intel_outreach(profile: dict) -> None:
    profile["outreach"] = {
        "cold_email_subject": "",
        "cold_email_body": "",
        "whatsapp_or_linkedin_message": "",
        "follow_up_email": "",
        "outreach_note": INTEL_ONLY_NOTICE,
    }


PLACEHOLDER_REPLACEMENTS = {
    "{{sender.name}}": SENDER_NAME,
    "{{sender.company}}": SENDER_COMPANY,
    "{{sender.title}}": SENDER_TITLE,
    "{{sender.email_signature}}": FIXED_SENDER_SIGNATURE,
    "{{sender.phone}}": "",
    "[Your Name]": SENDER_NAME,
    "[您的姓名]": SENDER_NAME,
    "[Your Position]": SENDER_TITLE,
    "[您的职位]": SENDER_TITLE,
    "[Your Title]": SENDER_TITLE,
    "[Your Company Name]": SENDER_COMPANY,
    "[您的公司名称]": SENDER_COMPANY,
    "[Your Company]": SENDER_COMPANY,
    "[您的公司]": SENDER_COMPANY,
    "[Your Phone]": "",
    "[Your Phone Number]": "",
    "[您的电话]": "",
    "[您的电话号码]": "",
}


def _normalize_sender_placeholders(text: str) -> str:
    out = text or ""
    for old, new in PLACEHOLDER_REPLACEMENTS.items():
        out = out.replace(old, new)
    return out


def _ensure_fixed_sender_signature(text: str) -> str:
    out = _normalize_sender_placeholders(_cell_text(text))
    if not out:
        return ""
    if FIXED_SENDER_SIGNATURE in out:
        return out
    return out.rstrip() + "\n\n" + FIXED_SENDER_SIGNATURE


def _normalize_outreach_templates(profile: dict) -> None:
    outreach = profile.setdefault("outreach", {})
    for key in OUTREACH_DRAFT_KEYS:
        outreach[key] = _normalize_sender_placeholders(_cell_text(outreach.get(key)))
    for key in ("cold_email_body", "follow_up_email"):
        outreach[key] = _ensure_fixed_sender_signature(outreach.get(key))


def _contains_cjk(text: str) -> bool:
    return bool(CJK_RE.search(text or ""))


def _outreach_cjk_fields(profile: dict) -> list[str]:
    outreach = profile.get("outreach") or {}
    return [key for key in OUTREACH_DRAFT_KEYS if _contains_cjk(_cell_text(outreach.get(key)))]


def _outreach_is_complete(profile: dict) -> bool:
    outreach = profile.get("outreach") or {}
    return all(_cell_text(outreach.get(key)) for key in OUTREACH_DRAFT_KEYS)


def _outreach_is_empty(profile: dict) -> bool:
    outreach = profile.get("outreach") or {}
    return not any(_cell_text(outreach.get(key)) for key in OUTREACH_DRAFT_KEYS)


def _coerce_list(value: object) -> list:
    if isinstance(value, list):
        return value
    if not value:
        return []
    return [value]


EVIDENCE_TYPES = {"identity", "business", "product_goji", "contact", "outreach_hook"}


def _guess_evidence_type(item: dict) -> str:
    text = " ".join(
        _cell_text(item.get(key))
        for key in ("evidence_type", "claim", "source_quote", "page_section")
    ).lower()
    if any(token in text for token in ("email", "phone", "tel", "contact", "address")):
        return "contact"
    if any(token in text for token in ("goji", "wolfberry", "lycium", "枸杞")):
        return "product_goji"
    if any(token in text for token in ("wholesale", "bulk", "distributor", "ingredient", "b2b", "manufacturer")):
        return "business"
    if any(token in text for token in ("founded", "based", "company", "family", "established")):
        return "identity"
    return "outreach_hook"


def _normalize_evidence_items(profile: dict, *, extracted_at: str) -> list[dict]:
    body = profile.setdefault("profile", {})
    evidence = _coerce_list(body.get("goji_or_adjacent_evidence"))
    normalized: list[dict] = []
    for raw in evidence:
        if not isinstance(raw, dict):
            continue
        evidence_type = _cell_text(raw.get("evidence_type"))
        if evidence_type not in EVIDENCE_TYPES:
            evidence_type = _guess_evidence_type(raw)
        normalized.append(
            {
                "evidence_type": evidence_type,
                "claim": _cell_text(raw.get("claim")),
                "source_url": _cell_text(raw.get("source_url")),
                "page_section": _cell_text(raw.get("page_section")),
                "source_quote": _cell_text(raw.get("source_quote")),
                "extracted_at": _cell_text(raw.get("extracted_at")) or extracted_at,
                "quote_supported": bool(raw.get("quote_supported", False)),
            }
        )
    body["goji_or_adjacent_evidence"] = normalized
    return normalized


CONTACT_SIGNAL_KEYS = (
    "emails",
    "phones",
    "phone_candidates_low_confidence",
    "whatsapp_links",
    "contact_pages",
    "social_links",
)


def _merge_contact_signals(profile: dict, contacts: ContactSignals) -> dict:
    existing = profile.get("contact_signals") if isinstance(profile.get("contact_signals"), dict) else {}
    merged: dict[str, list[str]] = {}
    for key in CONTACT_SIGNAL_KEYS:
        values: list[str] = []
        raw_existing = existing.get(key) if isinstance(existing, dict) else []
        if not isinstance(raw_existing, list):
            raw_existing = [raw_existing]
        values.extend(_cell_text(value) for value in raw_existing if _cell_text(value))
        values.extend(_cell_text(value) for value in getattr(contacts, key, []) if _cell_text(value))
        merged[key] = _unique(values)
    profile["contact_signals"] = merged
    return merged


def normalize_profile_draft(lead: VerifiedLead, profile: dict, contacts: ContactSignals) -> dict:
    if not isinstance(profile, dict):
        profile = {}
    generated_at = dt.datetime.now().isoformat(timespec="seconds")
    company = profile.setdefault("company", {})
    company.setdefault("display_name", lead.xiaoman_company_name or lead.input_company_name)
    company.setdefault("input_company_name", lead.input_company_name)
    company.setdefault("website", lead.website)
    company["country"] = _resolved_country_for_lead(lead)
    company.setdefault("step4_rating", lead.rating)
    company.setdefault("step4_customer_type", lead.customer_type)
    company.setdefault("step4_goji_presence", lead.goji_presence)

    body = profile.setdefault("profile", {})
    for key in (
        "bio_cn",
        "bio_en",
        "business_relevance_cn",
        "contact_summary_cn",
        "recommended_outreach_angle_cn",
        "internal_notes_cn",
    ):
        body[key] = _cell_text(body.get(key))
    body["missing_info_cn"] = [_cell_text(item) for item in _coerce_list(body.get("missing_info_cn")) if _cell_text(item)]
    body["outreach_risk_cn"] = [_cell_text(item) for item in _coerce_list(body.get("outreach_risk_cn")) if _cell_text(item)]

    evidence = _normalize_evidence_items(profile, extracted_at=generated_at)
    outreach_mode = _outreach_mode_for_lead(lead)
    if outreach_mode == "intel_only":
        _blank_intel_outreach(profile)
        if INTEL_ONLY_NOTICE not in body["outreach_risk_cn"]:
            body["outreach_risk_cn"].append(INTEL_ONLY_NOTICE)
    else:
        _normalize_outreach_templates(profile)

    profile["profile_id"] = f"{lead.row_number:03d}-{_slugify(lead.xiaoman_company_name or lead.input_company_name or lead.website)}"
    profile["profile_version"] = PROFILE_VERSION
    profile["generated_at"] = generated_at
    profile["outreach_mode"] = outreach_mode
    profile["evidence_insufficient"] = len(evidence) < MIN_PROFILE_EVIDENCE
    profile["sales_ready"] = False
    profile["eval_flags"] = []
    profile["quality_flags"] = []
    _merge_contact_signals(profile, contacts)
    profile["contact_quality"] = {
        "high_confidence_phone_count": len(contacts.phones),
        "low_confidence_phone_candidate_count": len(contacts.phone_candidates_low_confidence),
    }
    return profile


def _normalize_for_quote(text: str) -> str:
    return re.sub(r"\s+", " ", text or "").strip().lower()


def fact_check_profile(profile: dict, pages: list[PageSnapshot]) -> dict:
    corpus = _normalize_for_quote("\n".join(page.text for page in pages))
    evidence = (((profile.get("profile") or {}).get("goji_or_adjacent_evidence")) or [])
    checked = []
    unsupported = []
    checked_at = dt.datetime.now().isoformat(timespec="seconds")
    for item in evidence:
        quote = _cell_text(item.get("source_quote") if isinstance(item, dict) else "")
        quote_is_claimed = bool(quote) and quote != "未找到证据"
        supported = quote_is_claimed and _normalize_for_quote(quote) in corpus
        record = {
            "evidence_type": item.get("evidence_type", "") if isinstance(item, dict) else "",
            "claim": item.get("claim", "") if isinstance(item, dict) else "",
            "source_url": item.get("source_url", "") if isinstance(item, dict) else "",
            "page_section": item.get("page_section", "") if isinstance(item, dict) else "",
            "source_quote": quote,
            "extracted_at": item.get("extracted_at", checked_at) if isinstance(item, dict) else checked_at,
            "quote_supported": supported,
        }
        checked.append(record)
        if quote_is_claimed and not supported:
            unsupported.append(record)
    return {
        "checked_at": checked_at,
        "evidence_count": len(checked),
        "unsupported_quote_count": len(unsupported),
        "evidence": checked,
    }


def _apply_fact_check_to_profile(profile: dict, fact_check: dict) -> None:
    body = profile.setdefault("profile", {})
    checked = fact_check.get("evidence") or []
    body["goji_or_adjacent_evidence"] = checked


def _sender_profile_exists(verified_xlsx: Path, output_dir: Path) -> bool:
    candidates = [
        output_dir / "sender_profile.json",
        verified_xlsx.parent / "sender_profile.json",
        Path("sender_profile.json"),
    ]
    return any(path.is_file() and path.read_text(encoding="utf-8").strip() for path in candidates)


def _outreach_has_content(profile: dict) -> bool:
    outreach = profile.get("outreach") or {}
    return any(_cell_text(outreach.get(key)) for key in OUTREACH_DRAFT_KEYS)


def _supported_evidence_for_outreach(fact_check: dict) -> dict | None:
    for item in fact_check.get("evidence") or []:
        if item.get("quote_supported") and _cell_text(item.get("claim")):
            return item
    for item in fact_check.get("evidence") or []:
        if _cell_text(item.get("claim")):
            return item
    return None


def _trim_sentence(value: str, max_chars: int = 180) -> str:
    text = re.sub(r"\s+", " ", value or "").strip().strip("\"'")
    if len(text) <= max_chars:
        return text
    trimmed = text[:max_chars].rsplit(" ", 1)[0].strip()
    return trimmed.rstrip(".,;:") + "..."


def _outreach_evidence_items(fact_check: dict) -> list[dict]:
    evidence = [item for item in (fact_check.get("evidence") or []) if isinstance(item, dict)]
    priority = {
        "business": 0,
        "outreach_hook": 1,
        "product_goji": 2,
        "identity": 3,
        "contact": 4,
    }
    return sorted(evidence, key=lambda item: priority.get(_cell_text(item.get("evidence_type")), 9))


def _infer_business_track(profile: dict) -> str:
    body = profile.get("profile") if isinstance(profile.get("profile"), dict) else {}
    text = " ".join(
        [
            _cell_text(body.get("bio_cn")),
            _cell_text(body.get("bio_en")),
            _cell_text(body.get("business_relevance_cn")),
            _cell_text(body.get("recommended_outreach_angle_cn")),
        ]
    ).lower()
    for track, keywords in TRACK_KEYWORDS.items():
        if any(keyword.lower() in text for keyword in keywords):
            return track
    return TRACK_FALLBACK


def _supplier_angle(supplier_profile: dict, track: str) -> str:
    use_case_angles = supplier_profile.get("use_case_angles")
    if not isinstance(use_case_angles, dict):
        return ""
    return _cell_text(use_case_angles.get(track)) or _cell_text(use_case_angles.get(TRACK_FALLBACK))


def _supplier_certifications_for_country(lead: VerifiedLead, supplier_profile: dict) -> list[str]:
    supplier_certs = [
        _cell_text(value)
        for value in (supplier_profile.get("certifications") or [])
        if _cell_text(value)
    ]
    cert_lookup = {value.lower(): value for value in supplier_certs}
    country = " ".join([lead.input_country, lead.domain, lead.website]).lower()
    if any(
        token in country
        for token in (
            "europe",
            "germany",
            "france",
            "italy",
            "spain",
            "netherlands",
            "belgium",
            "austria",
            "switzerland",
            "poland",
            "sweden",
            "denmark",
            "norway",
            "finland",
            "united kingdom",
            "uk",
            ".de",
            ".fr",
            ".it",
            ".es",
            ".nl",
            ".eu",
        )
    ):
        preferred = ("EU BCS Organic", "ISO 22000")
    elif any(token in country for token in ("united states", "usa", "u.s.", "north america", "canada")) or re.search(
        r"\bus\b", country
    ):
        preferred = ("FDA", "KOSHER", "ISO 22000")
    elif any(
        token in country
        for token in (
            "middle east",
            "uae",
            "united arab emirates",
            "saudi",
            "qatar",
            "kuwait",
            "bahrain",
            "oman",
            "jordan",
            "israel",
            "turkey",
        )
    ):
        preferred = ("HALAL", "ISO 22000")
    else:
        preferred = ("FDA", "HALAL", "KOSHER", "ISO 22000")

    selected = [cert_lookup[name.lower()] for name in preferred if name.lower() in cert_lookup]
    return selected or supplier_certs[:2]


def _english_evidence_sentence(fact_check: dict, profile: dict | None = None) -> str:
    track_label = TRACK_LABELS.get(_infer_business_track(profile or {}), TRACK_LABELS[TRACK_FALLBACK])
    return (
        f"Your team appears aligned with {track_label} categories, so I am reaching out "
        "about Ningxia goji ingredient sourcing."
    )


def _ensure_email_references_evidence(profile: dict, fact_check: dict) -> None:
    # Profile evidence is retained for internal review; sales copy must not expose it.
    return


def _outreach_has_web_reference_language(profile: dict) -> bool:
    outreach = profile.get("outreach") or {}
    text = "\n".join(_cell_text(outreach.get(key)) for key in OUTREACH_DRAFT_KEYS)
    return bool(WEB_REFERENCE_OUTREACH_RE.search(text))


def _outreach_missing_required_sales_elements(profile: dict) -> bool:
    outreach = profile.get("outreach") or {}
    body = _cell_text(outreach.get("cold_email_body"))
    if not body:
        return False
    body_lower = body.lower()
    missing_fact = any(fact.lower() not in body_lower for fact in REQUIRED_SALES_EMAIL_FACTS)
    missing_cta = not any(hint in body_lower for hint in REQUIRED_SALES_CTA_HINTS)
    return missing_fact or missing_cta


def _safe_sales_outreach_template(
    lead: VerifiedLead,
    profile: dict,
    fact_check: dict,
    supplier_profile: dict,
) -> dict:
    raw_display_name = (
        (profile.get("company") or {}).get("display_name")
        or lead.xiaoman_company_name
        or lead.input_company_name
        or "there"
    )
    display_name = _cell_text(raw_display_name)
    if _contains_cjk(display_name):
        parsed_domain = urlparse(lead.website).netloc.replace("www.", "") if lead.website else ""
        display_name = parsed_domain or "there"
    track = _infer_business_track(profile)
    track_label = TRACK_LABELS.get(track, TRACK_LABELS[TRACK_FALLBACK])
    partner_label = TRACK_PARTNER_LABELS.get(track, TRACK_PARTNER_LABELS["other"])
    supplier_angle = _supplier_angle(supplier_profile, track)
    certs = _supplier_certifications_for_country(lead, supplier_profile)
    cert_text = ", ".join(certs) if certs else "FDA, HALAL, KOSHER, and ISO 22000"
    origin = _cell_text(supplier_profile.get("origin")) or "Ningxia, China"
    founded_year = _cell_text(supplier_profile.get("founded_year")) or "2003"
    subject = f"Bairuiyuan Ningxia goji for {display_name} {track_label}"
    body = (
        f"Dear {display_name} team,\n\n"
        "I am reaching out to ask whether your team is currently sourcing goji berries "
        "or goji extract for related product lines.\n\n"
        f"Bairuiyuan supplies Ningxia goji ingredients for {partner_label} partners. "
        f"For {track_label} programs, our goji can support {supplier_angle}.\n\n"
        f"Bairuiyuan was founded in {founded_year} in {origin}. We operate self-managed organic "
        "goji plantations and supply dried goji berries and goji extract, including LBP 10%-50%, "
        f"with relevant certifications such as {cert_text}.\n\n"
        "If this category is relevant, please reply with any target volume or spec needs, "
        "or connect me with the sourcing or product contact responsible for botanical ingredients. "
        "I would also be glad to arrange a 15-minute video call.\n\n"
        f"{FIXED_SENDER_SIGNATURE}"
    )
    message = (
        f"Hello {display_name} team, this is Nicky with Bairuiyuan Goji. We supply Ningxia "
        f"goji berries and extract for {track_label} and related botanical ingredient needs. "
        "Is your team reviewing goji sourcing, or could you connect me with the sourcing "
        "or product contact?"
    )
    follow_up = (
        f"Dear {display_name} team,\n\n"
        f"Following up on my note about {subject}. I can send a spec sheet or sample options "
        f"for Bairuiyuan Ningxia goji, including {cert_text}. Would it be useful to compare "
        "specs, or schedule a 15-minute video call?\n\n"
        f"{FIXED_SENDER_SIGNATURE}"
    )
    return {
        "cold_email_subject": subject,
        "cold_email_body": body,
        "whatsapp_or_linkedin_message": message,
        "follow_up_email": follow_up,
    }


def _enforce_english_sales_outreach(
    lead: VerifiedLead,
    profile: dict,
    fact_check: dict,
    supplier_profile: dict,
) -> bool:
    if profile.get("outreach_mode") != "sales":
        return False
    if not _outreach_cjk_fields(profile):
        return False
    profile["outreach"] = _safe_sales_outreach_template(lead, profile, fact_check, supplier_profile)
    profile["outreach"]["outreach_note"] = "English outreach template applied because the LLM draft contained Chinese text."
    return True


def _enforce_sales_outreach_no_web_reference_language(
    lead: VerifiedLead,
    profile: dict,
    fact_check: dict,
    supplier_profile: dict,
) -> bool:
    if profile.get("outreach_mode") != "sales":
        return False
    if not _outreach_has_web_reference_language(profile):
        return False
    profile["outreach"] = _safe_sales_outreach_template(lead, profile, fact_check, supplier_profile)
    profile["outreach"]["outreach_note"] = (
        "Safe Bairuiyuan outreach template applied because the LLM draft contained website-reference phrasing."
    )
    _normalize_outreach_templates(profile)
    return True


def _enforce_sales_outreach_required_elements(
    lead: VerifiedLead,
    profile: dict,
    fact_check: dict,
    supplier_profile: dict,
) -> bool:
    if profile.get("outreach_mode") != "sales":
        return False
    if not _outreach_missing_required_sales_elements(profile):
        return False
    profile["outreach"] = _safe_sales_outreach_template(lead, profile, fact_check, supplier_profile)
    profile["outreach"]["outreach_note"] = (
        "Safe Bairuiyuan outreach template applied because the LLM draft missed required supplier facts or CTA."
    )
    _normalize_outreach_templates(profile)
    return True


FORBIDDEN_OUTREACH_PATTERNS = re.compile(
    r"\b(USDA certified|competitive prices?|"
    r"production capacity|existing relationship|we(?:'| ha)ve helped)\b|"
    r"(竞争性价格|批发价格|稳定供应|供应能力|既有关系|成功添加)",
    re.IGNORECASE,
)


def _outreach_has_forbidden_claims(profile: dict) -> bool:
    outreach = profile.get("outreach") or {}
    text = "\n".join(_cell_text(outreach.get(key)) for key in outreach)
    return bool(FORBIDDEN_OUTREACH_PATTERNS.search(text))


def finalize_profile_quality(
    lead: VerifiedLead,
    profile: dict,
    fact_check: dict,
    *,
    sender_profile_present: bool,
    supplier_profile: dict,
) -> None:
    _apply_fact_check_to_profile(profile, fact_check)
    outreach_safe_template_applied = False
    llm_enrichment_failed = False
    if profile.get("outreach_mode") == "sales" and _outreach_is_empty(profile):
        profile["outreach"] = _safe_sales_outreach_template(lead, profile, fact_check, supplier_profile)
        llm_enrichment_failed = True
        outreach_safe_template_applied = True
        profile["outreach"]["outreach_note"] = (
            "Safe Bairuiyuan outreach template applied because the LLM draft returned empty outreach fields."
        )
    elif profile.get("outreach_mode") == "sales" and not sender_profile_present:
        profile["outreach"] = _safe_sales_outreach_template(lead, profile, fact_check, supplier_profile)
        outreach_safe_template_applied = True
    else:
        _ensure_email_references_evidence(profile, fact_check)
    _normalize_outreach_templates(profile)
    outreach_rewritten_en = _enforce_english_sales_outreach(lead, profile, fact_check, supplier_profile)
    if outreach_rewritten_en:
        outreach_safe_template_applied = True
    outreach_rewritten_web_reference = _enforce_sales_outreach_no_web_reference_language(
        lead,
        profile,
        fact_check,
        supplier_profile,
    )
    if outreach_rewritten_web_reference:
        outreach_safe_template_applied = True
    outreach_required_elements_rewritten = _enforce_sales_outreach_required_elements(
        lead,
        profile,
        fact_check,
        supplier_profile,
    )
    if outreach_required_elements_rewritten:
        outreach_safe_template_applied = True
    evidence_count = int(fact_check.get("evidence_count") or 0)
    unsupported_count = int(fact_check.get("unsupported_quote_count") or 0)
    evidence_insufficient = evidence_count < MIN_PROFILE_EVIDENCE
    outreach_cjk_fields = _outreach_cjk_fields(profile)
    outreach_complete = _outreach_is_complete(profile) if profile.get("outreach_mode") == "sales" else False

    flags: list[str] = []
    quality_flags: list[str] = []
    if profile.get("outreach_mode") == "intel_only":
        flags.append("intel_only")
    if llm_enrichment_failed:
        quality_flags.append("llm_enrichment_failed")
    if outreach_safe_template_applied:
        quality_flags.append("outreach_safe_template_applied")
    if outreach_rewritten_web_reference:
        quality_flags.append("outreach_web_reference_rewritten")
    if outreach_required_elements_rewritten:
        quality_flags.append("outreach_required_elements_rewritten")
    if outreach_rewritten_en:
        flags.append("outreach_language_rewritten_en")
    if outreach_cjk_fields:
        flags.append("outreach_contains_cjk")
    if profile.get("outreach_mode") == "sales" and not outreach_complete:
        flags.append("outreach_missing")
    if evidence_insufficient:
        flags.append("evidence_insufficient")
    if unsupported_count:
        flags.append("unsupported_quote")
    if not sender_profile_present:
        flags.append("sender_profile_missing")
    if profile.get("outreach_mode") == "intel_only" and _outreach_has_content(profile):
        flags.append("competitor_or_non_target_has_outreach")
    if _outreach_has_forbidden_claims(profile):
        flags.append("forbidden_outreach_claim")

    sales_ready = (
        profile.get("outreach_mode") == "sales"
        and not evidence_insufficient
        and unsupported_count == 0
        and sender_profile_present
        and outreach_complete
        and not outreach_cjk_fields
        and not _outreach_has_forbidden_claims(profile)
    )
    profile["evidence_insufficient"] = evidence_insufficient
    profile["sales_ready"] = sales_ready
    profile["eval_flags"] = flags
    profile["quality_flags"] = quality_flags
    profile["quality"] = {
        "outreach_mode": profile.get("outreach_mode"),
        "sales_ready": sales_ready,
        "evidence_count": evidence_count,
        "evidence_insufficient": evidence_insufficient,
        "unsupported_quote_count": unsupported_count,
        "sender_profile_present": sender_profile_present,
        "outreach_language": "en" if profile.get("outreach_mode") == "sales" else "",
        "outreach_language_rewritten": outreach_rewritten_en,
        "outreach_web_reference_rewritten": outreach_rewritten_web_reference,
        "outreach_required_elements_rewritten": outreach_required_elements_rewritten,
        "outreach_complete": outreach_complete,
        "outreach_cjk_fields": outreach_cjk_fields,
        "eval_flags": flags,
        "quality_flags": quality_flags,
        "step4_verified_target": lead.verified_target,
        "step4_is_competitor": lead.is_competitor,
    }


def render_profile_markdown(lead: VerifiedLead, contacts: ContactSignals, profile: dict, fact_check: dict) -> str:
    company = profile.get("company") or {}
    body = profile.get("profile") or {}
    outreach = profile.get("outreach") or {}
    evidence = body.get("goji_or_adjacent_evidence") or []

    lines = [
        f"# {company.get('display_name') or lead.xiaoman_company_name or lead.input_company_name}",
        "",
        "## Quality",
        "",
        f"- Profile Version: {profile.get('profile_version', PROFILE_VERSION)}",
        f"- Outreach Mode: {profile.get('outreach_mode', '')}",
        f"- Outreach Language: {(profile.get('quality') or {}).get('outreach_language', '')}",
        f"- Sales Ready: {profile.get('sales_ready', False)}",
        f"- Evidence Insufficient: {profile.get('evidence_insufficient', False)}",
        f"- Eval Flags: {'; '.join(profile.get('eval_flags') or []) or '无'}",
        "",
        "## Step4 Input",
        "",
        f"- Input Company: {lead.input_company_name}",
        f"- Xiaoman Company: {lead.xiaoman_company_name}",
        f"- Website: {lead.website}",
        f"- Country: {lead.input_country}",
        f"- Rating: {lead.rating}",
        f"- Customer Type: {lead.customer_type}",
        f"- Goji Presence: {lead.goji_presence}",
        f"- Is Competitor: {lead.is_competitor}",
        f"- Rating Reason: {lead.rating_reason}",
        f"- Outreach Angle: {lead.outreach_angle}",
        "",
        "## Profile",
        "",
        f"**公司简介**：{body.get('bio_cn', '')}",
        "",
        f"**Business Bio**: {body.get('bio_en', '')}",
        "",
        f"**业务相关性**：{body.get('business_relevance_cn', '')}",
        "",
        f"**联系方式摘要**：{body.get('contact_summary_cn', '')}",
        "",
        f"**推荐触达角度**：{body.get('recommended_outreach_angle_cn', '')}",
        "",
        "## Contact Signals",
        "",
        f"- Emails: {'; '.join(contacts.emails) or '未找到'}",
        f"- High-confidence Phones: {'; '.join(contacts.phones) or '未找到'}",
        f"- Low-confidence Phone-like Text (not for dialing): {'; '.join(contacts.phone_candidates_low_confidence) or '无'}",
        f"- WhatsApp: {'; '.join(contacts.whatsapp_links) or '未找到'}",
        f"- Contact Pages: {'; '.join(contacts.contact_pages) or '未找到'}",
        f"- Social Links: {'; '.join(contacts.social_links) or '未找到'}",
        "",
        "## Evidence",
        "",
    ]
    if evidence:
        for item in evidence:
            lines.extend(
                [
                    f"- Type: {item.get('evidence_type', '')}",
                    f"  Claim: {item.get('claim', '')}",
                    f"  Source: {item.get('source_url', '')}",
                    f"  Section: {item.get('page_section', '')}",
                    f"  Quote: {item.get('source_quote', '')}",
                    f"  Extracted At: {item.get('extracted_at', '')}",
                    f"  Quote Supported: {item.get('quote_supported', False)}",
                ]
            )
    else:
        lines.append("- 未找到证据")

    missing = body.get("missing_info_cn") or []
    risks = body.get("outreach_risk_cn") or []
    lines.extend(
        [
            "",
            "## Missing Info / Risk",
            "",
            f"- Missing Info: {'; '.join(missing) if missing else '无'}",
            f"- Outreach Risk: {'; '.join(risks) if risks else '无'}",
            f"- Internal Notes: {body.get('internal_notes_cn', '')}",
            f"- Unsupported Evidence Quotes: {fact_check.get('unsupported_quote_count', 0)}",
            "",
            "## Outreach Drafts",
            "",
            f"- Outreach Note: {outreach.get('outreach_note', '') or ('可用于销售模板复核' if profile.get('outreach_mode') == 'sales' else '')}",
            "",
            f"### Subject\n{outreach.get('cold_email_subject', '')}",
            "",
            f"### Cold Email\n{outreach.get('cold_email_body', '')}",
            "",
            f"### WhatsApp / LinkedIn\n{outreach.get('whatsapp_or_linkedin_message', '')}",
            "",
            f"### Follow-up\n{outreach.get('follow_up_email', '')}",
            "",
        ]
    )
    return "\n".join(lines).rstrip() + "\n"


def _write_json(path: Path, data: object) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")


def _read_json_optional(path: Path) -> object | None:
    try:
        if path.is_file():
            return json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        return None
    return None


def _read_cached_pages(path: Path) -> list[PageSnapshot] | None:
    data = _read_json_optional(path)
    if not isinstance(data, list):
        return None
    pages: list[PageSnapshot] = []
    for item in data:
        if not isinstance(item, dict):
            continue
        url = _cell_text(item.get("url"))
        text = _cell_text(item.get("text"))
        if not url or not text:
            continue
        pages.append(
            PageSnapshot(
                url=url,
                status_code=int(item.get("status_code") or 0),
                title=_cell_text(item.get("title")),
                text=text,
                tel_links=[
                    _cell_text(value)
                    for value in (item.get("tel_links") or [])
                    if _cell_text(value)
                ],
                crawler=_cell_text(item.get("crawler")) or "cache",
            )
        )
    return pages or None


def _read_cached_contacts(path: Path) -> ContactSignals | None:
    data = _read_json_optional(path)
    if not isinstance(data, dict):
        return None
    return ContactSignals(
        emails=[_cell_text(value) for value in (data.get("emails") or []) if _cell_text(value)],
        phones=[_cell_text(value) for value in (data.get("phones") or []) if _cell_text(value)],
        phone_candidates_low_confidence=[
            _cell_text(value)
            for value in (data.get("phone_candidates_low_confidence") or [])
            if _cell_text(value)
        ],
        whatsapp_links=[
            _cell_text(value)
            for value in (data.get("whatsapp_links") or [])
            if _cell_text(value)
        ],
        contact_pages=[
            _cell_text(value)
            for value in (data.get("contact_pages") or [])
            if _cell_text(value)
        ],
        social_links=[
            _cell_text(value)
            for value in (data.get("social_links") or [])
            if _cell_text(value)
        ],
    )


def _ensure_feedback_files(output_dir: Path) -> None:
    feedback_path = output_dir / "feedback.jsonl"
    feedback_path.parent.mkdir(parents=True, exist_ok=True)
    feedback_path.touch(exist_ok=True)
    _write_json(
        output_dir / "feedback_schema.json",
        {
            "format": "jsonl",
            "fields": FEEDBACK_SCHEMA_FIELDS,
            "template": {field_name: "" for field_name in FEEDBACK_SCHEMA_FIELDS},
        },
    )


def _contact_low_conf_count(contacts: ContactSignals | dict | None) -> int:
    if isinstance(contacts, ContactSignals):
        return len(contacts.phone_candidates_low_confidence)
    if isinstance(contacts, dict):
        return len(contacts.get("phone_candidates_low_confidence") or [])
    return 0


def _record_quality_fields(
    lead: VerifiedLead,
    profile: dict | None,
    fact_check: dict | None,
    contacts: ContactSignals | dict | None,
) -> dict:
    profile = profile or {}
    fact_check = fact_check or {}
    evidence_count = int(fact_check.get("evidence_count") or 0)
    if not evidence_count:
        evidence_count = len(((profile.get("profile") or {}).get("goji_or_adjacent_evidence")) or [])
    outreach_mode = _cell_text(profile.get("outreach_mode")) or _outreach_mode_for_lead(lead)
    evidence_insufficient = bool(profile.get("evidence_insufficient", evidence_count < MIN_PROFILE_EVIDENCE))
    unsupported_quote_count = int(fact_check.get("unsupported_quote_count") or 0)
    competitor_with_outreach = (
        outreach_mode == "intel_only" or _is_yes(lead.is_competitor) or lead.customer_type == "竞争对手"
    ) and _outreach_has_content(profile)
    eval_flags = profile.get("eval_flags") or []
    quality_flags = profile.get("quality_flags") or (profile.get("quality") or {}).get("quality_flags") or []
    return {
        "profile_version": _cell_text(profile.get("profile_version")) or PROFILE_VERSION,
        "outreach_mode": outreach_mode,
        "sales_ready": bool(profile.get("sales_ready", False)),
        "evidence_count": evidence_count,
        "evidence_insufficient": evidence_insufficient,
        "unsupported_quote_count": unsupported_quote_count,
        "eval_flags": eval_flags,
        "quality_flags": quality_flags,
        "outreach_language_rewritten": "outreach_language_rewritten_en" in eval_flags,
        "outreach_contains_cjk": "outreach_contains_cjk" in eval_flags,
        "outreach_missing": "outreach_missing" in eval_flags,
        "competitor_with_outreach": competitor_with_outreach,
        "dirty_phone_candidate_count": _contact_low_conf_count(contacts),
    }


def build_eval_summary(index_records: list[dict]) -> dict:
    profile_records = [record for record in index_records if record.get("status") == "profile_written"]
    return {
        "generated_at": dt.datetime.now().isoformat(timespec="seconds"),
        "profile_version": PROFILE_VERSION,
        "total_profiles": len(profile_records),
        "sales_ready_count": sum(1 for record in profile_records if record.get("sales_ready")),
        "intel_only_count": sum(1 for record in profile_records if record.get("outreach_mode") == "intel_only"),
        "profiles_with_evidence_gte_3": sum(
            1 for record in profile_records if int(record.get("evidence_count") or 0) >= MIN_PROFILE_EVIDENCE
        ),
        "evidence_insufficient_count": sum(
            1 for record in profile_records if record.get("evidence_insufficient")
        ),
        "unsupported_quote_total": sum(
            int(record.get("unsupported_quote_count") or 0) for record in profile_records
        ),
        "competitor_with_outreach_count": sum(
            1 for record in profile_records if record.get("competitor_with_outreach")
        ),
        "outreach_language_rewritten_count": sum(
            1 for record in profile_records if record.get("outreach_language_rewritten")
        ),
        "outreach_contains_cjk_count": sum(
            1 for record in profile_records if record.get("outreach_contains_cjk")
        ),
        "outreach_missing_count": sum(
            1 for record in profile_records if record.get("outreach_missing")
        ),
        "dirty_phone_candidate_count": sum(
            int(record.get("dirty_phone_candidate_count") or 0) for record in profile_records
        ),
    }


def _sales_lookup_key(value: object) -> str:
    return re.sub(r"\s+", " ", _cell_text(value).lower()).strip()


def _sales_lookup_keys(lead: VerifiedLead, profile: dict | None = None) -> list[tuple[str, str]]:
    profile = profile or {}
    company = profile.get("company") if isinstance(profile.get("company"), dict) else {}
    keys: list[tuple[str, str]] = []
    for value in (
        company.get("display_name"),
        company.get("input_company_name"),
        lead.xiaoman_company_name,
        lead.input_company_name,
    ):
        key = _sales_lookup_key(value)
        if key:
            keys.append(("name", key))
    for value in (company.get("website"), lead.website, lead.domain):
        key = _base_url(_cell_text(value)).lower()
        if key:
            keys.append(("site", key))
    return keys


def _contacts_list(contacts: object | None, key: str) -> list[str]:
    if isinstance(contacts, ContactSignals):
        values = getattr(contacts, key, [])
    elif isinstance(contacts, dict):
        values = contacts.get(key) or []
    else:
        values = []
    if not isinstance(values, list):
        values = [values]
    return [_cell_text(value) for value in values if _cell_text(value)]


def _sales_email_for_lead(lead: VerifiedLead, contacts: object | None) -> str:
    lead_email = _cell_text(lead.email)
    if lead_email:
        return lead_email
    emails = _contacts_list(contacts, "emails")
    return emails[0] if emails else ""


def _sales_outreach_text(outreach: dict, key: str) -> str:
    return _normalize_sender_placeholders(_cell_text(outreach.get(key)))


def build_sales_lead_rows(leads: list[VerifiedLead], output_dir: Path) -> list[dict]:
    profiles_by_slug: dict[str, dict] = {}
    contacts_by_slug: dict[str, object] = {}
    profiles_by_key: dict[tuple[str, str], dict] = {}
    contacts_by_key: dict[tuple[str, str], object] = {}

    for lead in leads:
        slug = _lead_slug(lead)
        profile = _read_json_optional(output_dir / "profiles" / f"{slug}.json")
        contacts = _read_json_optional(output_dir / "contacts" / f"{slug}.json")
        if isinstance(profile, dict):
            profiles_by_slug[slug] = profile
            for key in _sales_lookup_keys(lead, profile):
                profiles_by_key.setdefault(key, profile)
        if isinstance(contacts, dict):
            contacts_by_slug[slug] = contacts
            for key in _sales_lookup_keys(lead, profile if isinstance(profile, dict) else None):
                contacts_by_key.setdefault(key, contacts)

    rows: list[dict] = []
    for lead in leads:
        slug = _lead_slug(lead)
        profile = profiles_by_slug.get(slug)
        contacts = contacts_by_slug.get(slug)
        if profile is None:
            for key in _sales_lookup_keys(lead):
                profile = profiles_by_key.get(key)
                if profile is not None:
                    break
        if contacts is None:
            for key in _sales_lookup_keys(lead, profile):
                contacts = contacts_by_key.get(key)
                if contacts is not None:
                    break
        if not isinstance(profile, dict):
            continue
        if _cell_text(profile.get("outreach_mode")).lower() != "sales":
            continue

        email = _sales_email_for_lead(lead, contacts)
        phones = _contacts_list(contacts, "phones")
        contact_pages = _contacts_list(contacts, "contact_pages")
        if not (email or phones or contact_pages):
            continue
        can_email = bool(email)

        company = profile.get("company") if isinstance(profile.get("company"), dict) else {}
        body = profile.get("profile") if isinstance(profile.get("profile"), dict) else {}
        outreach = profile.get("outreach") if isinstance(profile.get("outreach"), dict) else {}
        resolved_country = _resolved_country_for_lead(lead) or _cell_text(company.get("country"))
        rows.append(
            {
                "公司名": _cell_text(company.get("display_name")) or lead.input_company_name,
                "国家": resolved_country,
                "网站": _cell_text(company.get("website")) or lead.website,
                "公司介绍": _cell_text(body.get("bio_cn")),
                "业务关联": _cell_text(body.get("business_relevance_cn")),
                "联系人": lead.contact_name,
                "职位": lead.position,
                "邮箱": email,
                "电话": "; ".join(phones),
                "有Contact Page": "Yes" if contact_pages else "",
                "邮件主题": _sales_outreach_text(outreach, "cold_email_subject") if can_email else "",
                "邮件正文": _sales_outreach_text(outreach, "cold_email_body") if can_email else "",
                "WhatsApp/LinkedIn": _sales_outreach_text(outreach, "whatsapp_or_linkedin_message"),
                "Follow-up 邮件": _sales_outreach_text(outreach, "follow_up_email") if can_email else "",
                "状态": "",
            }
        )
    return rows


def run(
    verified_xlsx: Path,
    output_dir: Path | None = None,
    *,
    sleep_s: float = 1.0,
    skip_llm: bool = False,
    dry_run: bool = False,
    overwrite: bool = False,
    limit: int | None = None,
) -> Path:
    if not verified_xlsx.is_file():
        raise FileNotFoundError(f"missing verified xlsx: {verified_xlsx}")
    supplier_profile = load_supplier_profile()
    output_dir = output_dir or (verified_xlsx.parent / "05_profiles")
    output_dir.mkdir(parents=True, exist_ok=True)
    _ensure_feedback_files(output_dir)
    sender_profile_present = True

    leads = load_enrich_leads(verified_xlsx)
    if limit is not None:
        leads = leads[: max(0, limit)]
    selected_path = output_dir / "selected_leads.json"
    _write_json(selected_path, [asdict(lead) for lead in leads])
    log.info("selected %d leads from %s", len(leads), verified_xlsx)
    if dry_run:
        for lead in leads:
            log.info(
                "selected row=%s rating=%s company=%s website=%s",
                lead.row_number,
                lead.rating,
                lead.xiaoman_company_name or lead.input_company_name,
                lead.website,
            )
        return output_dir

    index_records: list[dict] = []
    for idx, lead in enumerate(leads, start=1):
        slug = _lead_slug(lead)
        profile_json = output_dir / "profiles" / f"{slug}.json"
        profile_md = output_dir / "profiles" / f"{slug}.md"
        raw_json = output_dir / "raw_sites" / f"{slug}.json"
        raw_md = output_dir / "raw_sites" / f"{slug}.md"
        contacts_json = output_dir / "contacts" / f"{slug}.json"
        fact_json = output_dir / "fact_checks" / f"{slug}.json"
        error_json = output_dir / "errors" / f"{slug}.json"

        if profile_json.exists() and not overwrite:
            log.info("[%d/%d] %s already enriched, skipping", idx, len(leads), slug)
            existing_profile = _read_json_optional(profile_json)
            existing_fact_check = _read_json_optional(fact_json)
            existing_contacts = _read_json_optional(contacts_json)
            record = {
                "slug": slug,
                "row_number": lead.row_number,
                "company": lead.xiaoman_company_name or lead.input_company_name,
                "website": lead.website,
                "status": "profile_written",
                "profile_json": str(profile_json),
                "profile_md": str(profile_md),
                "fact_check": str(fact_json),
            }
            if isinstance(existing_profile, dict):
                record.update(
                    _record_quality_fields(
                        lead,
                        existing_profile,
                        existing_fact_check if isinstance(existing_fact_check, dict) else None,
                        existing_contacts if isinstance(existing_contacts, dict) else None,
                    )
                )
            if fact_json.exists():
                try:
                    fact_check = json.loads(fact_json.read_text(encoding="utf-8"))
                    record["unsupported_quote_count"] = fact_check.get("unsupported_quote_count", 0)
                except Exception:
                    record["unsupported_quote_count"] = ""
            index_records.append(record)
            continue

        log.info("[%d/%d] enriching %s", idx, len(leads), lead.xiaoman_company_name or lead.input_company_name)
        pages = _read_cached_pages(raw_json)
        if pages is None:
            pages = fetch_site_pages(lead.website)
            _write_json(raw_json, [asdict(page) for page in pages])
            raw_md.parent.mkdir(parents=True, exist_ok=True)
            raw_md.write_text(render_site_markdown(lead, pages), encoding="utf-8")
        else:
            log.info("[%d/%d] using cached website pages for %s", idx, len(leads), slug)

        contacts = _read_cached_contacts(contacts_json)
        if contacts is None:
            contacts = extract_contacts(pages, lead)
            _write_json(contacts_json, asdict(contacts))
        else:
            log.info("[%d/%d] using cached contact signals for %s", idx, len(leads), slug)

        record = {
            "slug": slug,
            "row_number": lead.row_number,
            "company": lead.xiaoman_company_name or lead.input_company_name,
            "website": lead.website,
            "pages_fetched": len(pages),
            "emails_found": len(contacts.emails),
            "phones_found": len(contacts.phones),
            "dirty_phone_candidate_count": len(contacts.phone_candidates_low_confidence),
            "outreach_mode": _outreach_mode_for_lead(lead),
            "raw_site": str(raw_md),
            "contacts": str(contacts_json),
        }
        if skip_llm:
            record["status"] = "crawled_skip_llm"
            index_records.append(record)
        else:
            try:
                raw_profile = synthesize_profile(lead, pages, contacts, supplier_profile)
                if not isinstance(raw_profile, dict):
                    raw_profile = {}
                if _outreach_mode_for_lead(lead) == "sales":
                    raw_outreach = raw_profile.get("outreach")
                    raw_profile["outreach"] = rewrite_outreach_tone(
                        raw_outreach if isinstance(raw_outreach, dict) else {}
                    )
                profile = normalize_profile_draft(lead, raw_profile, contacts)
                fact_check = fact_check_profile(profile, pages)
                finalize_profile_quality(
                    lead,
                    profile,
                    fact_check,
                    sender_profile_present=sender_profile_present,
                    supplier_profile=supplier_profile,
                )
                _write_json(profile_json, profile)
                _write_json(fact_json, fact_check)
                profile_md.parent.mkdir(parents=True, exist_ok=True)
                profile_md.write_text(render_profile_markdown(lead, contacts, profile, fact_check), encoding="utf-8")
                record.update(
                    {
                        "status": "profile_written",
                        "profile_json": str(profile_json),
                        "profile_md": str(profile_md),
                        "fact_check": str(fact_json),
                    }
                )
                record.update(_record_quality_fields(lead, profile, fact_check, contacts))
            except Exception as exc:
                error_record = {
                    "failed_at": dt.datetime.now().isoformat(timespec="seconds"),
                    "slug": slug,
                    "company": lead.xiaoman_company_name or lead.input_company_name,
                    "error": repr(exc),
                }
                _write_json(error_json, error_record)
                log.error("profile synthesis failed for %s: %s", slug, exc)
                record.update(
                    {
                        "status": "profile_failed",
                        "error": repr(exc),
                        "error_json": str(error_json),
                    }
                )
            index_records.append(record)

        _write_json(output_dir / "profiles_index.json", index_records)
        _write_json(output_dir / "05_eval_summary.json", build_eval_summary(index_records))
        if sleep_s > 0 and idx < len(leads):
            time.sleep(sleep_s)

    _write_json(output_dir / "profiles_index.json", index_records)
    _write_json(output_dir / "05_eval_summary.json", build_eval_summary(index_records))
    sales_lead_rows = build_sales_lead_rows(leads, output_dir)
    sales_leads_xlsx = output_dir / "05_sales_leads.xlsx"
    write_sales_leads_xlsx(sales_lead_rows, sales_leads_xlsx)
    log.info("profile enrich wrote index to %s", output_dir / "profiles_index.json")
    log.info("profile enrich wrote %d sales leads to %s", len(sales_lead_rows), sales_leads_xlsx)
    return output_dir


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Run step5 sales profile enrichment.")
    parser.add_argument("verified_xlsx", type=Path, help="Path to step4 04_verified.xlsx.")
    parser.add_argument(
        "--output-dir",
        type=Path,
        default=None,
        help="Output directory. Defaults to <run_dir>/05_profiles.",
    )
    parser.add_argument("--sleep", type=float, default=1.0, help="Seconds between companies.")
    parser.add_argument("--skip-llm", action="store_true", help="Fetch pages and contacts only.")
    parser.add_argument("--dry-run", action="store_true", help="Only list selected leads.")
    parser.add_argument("--overwrite", action="store_true", help="Regenerate existing profiles.")
    parser.add_argument("--limit", type=int, default=None, help="Process only the first N selected leads.")
    return parser.parse_args()


def main() -> None:
    load_dotenv()
    logging.basicConfig(level=logging.INFO, format="%(asctime)s  %(levelname)-7s %(message)s")
    args = parse_args()
    run(
        args.verified_xlsx,
        args.output_dir,
        sleep_s=args.sleep,
        skip_llm=args.skip_llm,
        dry_run=args.dry_run,
        overwrite=args.overwrite,
        limit=args.limit,
    )


if __name__ == "__main__":
    main()

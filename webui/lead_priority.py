"""Lead ranking: combine objective verdicts and sales context into a score."""
from __future__ import annotations

import json
import re
import sqlite3
from dataclasses import dataclass
from functools import lru_cache
from pathlib import Path
from typing import Optional

DEFAULT_PRIORITY = 0.5

STATUS_PRIORITY = {
    "已收到有效回复": 1.0,
    "已询价": 0.95,
    "已成交": 0.75,
    "已发": 0.5,
    "未发": 0.4,
    "联系人不对": 0.15,
    "不相关": 0.05,
    "国家错": 0.05,
    "邮箱无效": 0.05,
}


@dataclass
class RankingInputs:
    status: str = ""
    rating: str = ""
    customer_type: str = ""
    primary_vertical: str = ""
    food_supplement_focus: str = ""
    input_country: str = ""
    verdict_country: str = ""
    has_verified_email: bool = False
    feedback_score: Optional[float] = None


def priority_for_status(status: str | None) -> float:
    """Return the baseline priority for a lead status."""
    return STATUS_PRIORITY.get(str(status or "").strip(), DEFAULT_PRIORITY)


def compute_final_score(inp: RankingInputs) -> float:
    """Return final lead priority as a 0-1 float."""
    total = (
        _rating_score(inp.rating)
        + _status_bonus(inp.status)
        + _vertical_adj(inp.primary_vertical, inp.food_supplement_focus)
        + _country_adj(inp.input_country, inp.verdict_country)
        + _customer_bonus(inp.customer_type)
        + _email_adj(inp.has_verified_email)
        + _feedback_adj(inp.feedback_score)
    )
    return max(0.0, min(100.0, total)) / 100.0


def load_ranking_inputs(
    conn: sqlite3.Connection,
    runs_dir: Path,
    slug: str,
    run_id: str,
    *,
    status_override: str | None = None,
) -> RankingInputs:
    """Load ranking inputs from DB, step4 xlsx, profile JSON, and contacts JSON."""
    status = status_override or "未发"
    feedback_score: float | None = None
    try:
        status_row = conn.execute(
            """
            SELECT effective_status, feedback_score
            FROM lead_status
            WHERE profile_slug = ? AND run_id = ?
            """,
            (slug, run_id),
        ).fetchone()
    except sqlite3.OperationalError:
        status_row = conn.execute(
            """
            SELECT effective_status
            FROM lead_status
            WHERE profile_slug = ? AND run_id = ?
            """,
            (slug, run_id),
        ).fetchone()
    if status_row:
        status = status_override or str(status_row["effective_status"] or "未发")
        try:
            raw_feedback = status_row["feedback_score"]
            feedback_score = float(raw_feedback) if raw_feedback is not None else None
        except (IndexError, KeyError, TypeError, ValueError):
            feedback_score = None

    profile = _read_json(_profile_json_path(runs_dir, run_id, slug))
    contacts = _read_json(_contacts_json_path(runs_dir, run_id, slug))
    company = profile.get("company") if isinstance(profile.get("company"), dict) else {}
    contact_signals = (
        profile.get("contact_signals")
        if isinstance(profile.get("contact_signals"), dict)
        else {}
    )
    verified = _verified_row_for_slug(runs_dir, run_id, slug)

    contact_emails = _as_list(contacts.get("emails") if isinstance(contacts, dict) else [])
    profile_emails = _as_list(contact_signals.get("emails"))
    verified_email = str(verified.get("Email") or "").strip()
    has_verified_email = bool(contact_emails or profile_emails or verified_email)

    step4_customer_type = (
        str(company.get("step4_customer_type") or "")
        or str(verified.get("Customer Type") or "")
    )
    lead_type = str(verified.get("Lead Type") or "")
    customer_type = " | ".join(
        value
        for value in (step4_customer_type, lead_type)
        if value
    )

    return RankingInputs(
        status=status,
        rating=str(company.get("step4_rating") or verified.get("Rating") or ""),
        customer_type=customer_type,
        primary_vertical=str(
            company.get("step4_primary_vertical")
            or verified.get("Primary Vertical")
            or ""
        ),
        food_supplement_focus=str(
            company.get("step4_food_supplement_focus")
            or verified.get("Food/Supp Focus")
            or ""
        ),
        input_country=str(verified.get("Input Country") or ""),
        verdict_country=str(
            verified.get("Website Country")
            or company.get("country")
            or ""
        ),
        has_verified_email=has_verified_email,
        feedback_score=feedback_score,
    )


def _rating_score(rating: str) -> float:
    return {"S": 60.0, "A": 50.0, "P": 45.0, "B": 25.0, "C": 5.0, "Z": 0.0}.get(
        str(rating or "").strip().upper(),
        30.0,
    )


def _status_bonus(status: str) -> float:
    return {
        "已收到有效回复": 30.0,
        "已询价": 25.0,
        "已成交": 15.0,
        "已发": 0.0,
        "未发": -5.0,
        "联系人不对": -20.0,
        "不相关": -40.0,
        "国家错": -40.0,
        "邮箱无效": -40.0,
    }.get(str(status or "").strip(), 0.0)


def _vertical_adj(vertical: str, focus: str) -> float:
    vertical_key = str(vertical or "other").strip().lower()
    focus_key = str(focus or "none").strip().lower()

    if vertical_key in {"supplement", "food_beverage", "herbal_medicine"}:
        if focus_key == "core":
            score = 5.0
        elif focus_key == "partial":
            score = 0.0
        else:
            score = -10.0
    elif vertical_key == "cosmetic_beauty":
        score = 0.0 if focus_key in {"core", "partial"} else -10.0
    elif vertical_key in {"fitness_equipment", "general_marketplace"}:
        score = -20.0
    elif vertical_key in {"agriculture_raw", "other", ""}:
        score = -10.0
    else:
        score = -10.0

    if focus_key in {"marginal", "none", ""}:
        score = min(score, -10.0)
    return score


def _country_adj(input_country: str, verdict_country: str) -> float:
    input_norm = _normalize_country(input_country)
    verdict_norm = _normalize_country(verdict_country)
    if not input_norm or not verdict_norm or input_norm == verdict_norm:
        return 0.0
    return -15.0


def _customer_bonus(customer_type: str) -> float:
    text = str(customer_type or "").lower()
    if any(token in text for token in ("direct buyer", "oem制造商", "工厂客户", "品牌商", "品牌方")):
        return 8.0
    if any(token in text for token in ("distributor", "原料分销商", "分销商")):
        return 5.0
    if "retailer" in text or "零售" in text:
        return 3.0
    return 0.0


def _email_adj(has_verified_email: bool) -> float:
    return 0.0 if has_verified_email else -10.0


def _feedback_adj(feedback_score: Optional[float]) -> float:
    if feedback_score is None:
        return 0.0
    try:
        score = max(0.0, min(1.0, float(feedback_score)))
    except (TypeError, ValueError):
        return 0.0
    return (score - 0.5) * 30.0


_COUNTRY_ALIASES = {
    "ru": "russia",
    "rus": "russia",
    "russian federation": "russia",
    "russia": "russia",
    "tr": "turkey",
    "turkiye": "turkey",
    "türkiye": "turkey",
    "turkey": "turkey",
    "us": "united states",
    "usa": "united states",
    "u.s.": "united states",
    "u.s.a.": "united states",
    "united states of america": "united states",
    "uk": "united kingdom",
    "u.k.": "united kingdom",
    "great britain": "united kingdom",
}


def _normalize_country(value: str) -> str:
    text = re.sub(r"\s+", " ", str(value or "").strip().lower())
    text = text.replace("_", " ").replace("-", " ")
    return _COUNTRY_ALIASES.get(text, text)


_SAFE_COMPONENT_RE = re.compile(r"^[A-Za-z0-9][A-Za-z0-9._-]*$")


def _safe_component(value: str) -> bool:
    return bool(_SAFE_COMPONENT_RE.match(str(value or "")))


def _profile_json_path(runs_dir: Path, run_id: str, slug: str) -> Path:
    if not (_safe_component(run_id) and _safe_component(slug)):
        return Path()
    return runs_dir / run_id / "05_profiles" / "profiles" / f"{slug}.json"


def _contacts_json_path(runs_dir: Path, run_id: str, slug: str) -> Path:
    if not (_safe_component(run_id) and _safe_component(slug)):
        return Path()
    return runs_dir / run_id / "05_profiles" / "contacts" / f"{slug}.json"


def _read_json(path: Path) -> dict:
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return {}
    return data if isinstance(data, dict) else {}


def _as_list(value: object) -> list:
    if isinstance(value, list):
        return value
    if value:
        return [value]
    return []


def _verified_row_for_slug(runs_dir: Path, run_id: str, slug: str) -> dict[str, str]:
    if not _safe_component(run_id):
        return {}
    verified_xlsx = runs_dir / run_id / "04_verified.xlsx"
    if not verified_xlsx.is_file():
        return {}
    try:
        mtime_ns = verified_xlsx.stat().st_mtime_ns
    except OSError:
        return {}
    return _verified_rows_by_slug(str(verified_xlsx), mtime_ns).get(slug, {})


@lru_cache(maxsize=32)
def _verified_rows_by_slug(path_str: str, mtime_ns: int) -> dict[str, dict[str, str]]:
    del mtime_ns
    from openpyxl import load_workbook

    path = Path(path_str)
    rows_by_slug: dict[str, dict[str, str]] = {}
    try:
        workbook = load_workbook(path, read_only=True)
        sheet = workbook.active
        rows = sheet.iter_rows(values_only=True)
        header_row = next(rows, None)
        if not header_row:
            return rows_by_slug
        header = [str(value or "").strip() for value in header_row]
        for row_number, raw in enumerate(rows, start=2):
            row = {
                key: str(raw[idx] if idx < len(raw) and raw[idx] is not None else "").strip()
                for idx, key in enumerate(header)
            }
            if not any(row.values()):
                continue
            basis = (
                row.get("Xiaoman Company Name")
                or row.get("Input Company Name")
                or row.get("Website")
                or "company"
            )
            slug = f"{row_number:03d}-{_slugify(basis)}"
            rows_by_slug[slug] = row
    except OSError:
        return {}
    return rows_by_slug


def _slugify(value: str, fallback: str = "company") -> str:
    text = str(value or "").lower()
    text = re.sub(r"https?://", "", text)
    text = re.sub(r"[^a-z0-9]+", "-", text)
    text = text.strip("-")
    return (text or fallback)[:80]

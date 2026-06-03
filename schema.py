"""
Output xlsx schema for step2 → step3 bridge, and step3's own output.

step2 produces `02_buyers.xlsx` with BUYERS_COLUMNS. step3 reads that,
looks each company up on xiaoman, and produces `03_xiaoman.xlsx` with
XIAOMAN_COLUMNS — one row per buyer-match pair, except rank-1 rows may
expand into multiple rows when the matched company has multiple contacts.
"""
from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# Column order matters for human readability; step3 maps by header name not index.
BUYERS_COLUMNS = [
    "Company Name",
    "Country",
    "Lead Type",
    "Keywords Used",
    "Source Modifier",
    "Domain",
]


def write_buyers_xlsx(rows: list[dict], output_path: Path) -> None:
    """
    Write a buyers xlsx that step3's discovery-cli.js can consume directly.

    Each row dict should have all keys in BUYERS_COLUMNS. Missing keys become "".
    """
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Buyers"

    sheet.append(BUYERS_COLUMNS)
    for row in rows:
        sheet.append([row.get(col, "") for col in BUYERS_COLUMNS])

    sheet.freeze_panes = "A2"
    sheet.column_dimensions["A"].width = 40
    sheet.column_dimensions["B"].width = 18
    sheet.column_dimensions["C"].width = 16
    sheet.column_dimensions["D"].width = 44
    sheet.column_dimensions["E"].width = 18
    sheet.column_dimensions["F"].width = 28

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


XIAOMAN_COLUMNS = [
    "Input Company Name",
    "Input Country",
    "Input Lead Type",
    "Match Rank",
    "Xiaoman Company Name",
    "Xiaoman Country",
    "Xiaoman Country Code",
    "Domain",
    "Website",
    "Description",
    "Contact Count",
    "Email Count",
    "Company Hash ID",
    "Name Similarity",
    "Country Match",
    "Match Quality",
    "Top-1 Eligible",
    "Country Conflict",
    "Contact Name",
    "Position",
    "Email",
    "Email Quality",
    "Phone",
    "LinkedIn",
    "Confidence",
]


def write_xiaoman_xlsx(rows: list[dict], output_path: Path) -> None:
    """Write step3 output. Rows must have every key in XIAOMAN_COLUMNS."""
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Xiaoman"
    sheet.append(XIAOMAN_COLUMNS)
    for row in rows:
        sheet.append([row.get(col, "") for col in XIAOMAN_COLUMNS])
    sheet.freeze_panes = "A2"
    for col_letter, width in zip(
        "ABCDEFGHIJKLMNOPQRSTUVWXYZ"[: len(XIAOMAN_COLUMNS)],
        [32, 16, 14, 8, 32, 16, 6, 28, 32, 50, 8, 8, 34, 14, 12, 16, 12, 14, 28, 24, 40, 24, 24, 36, 12],
    ):
        sheet.column_dimensions[col_letter].width = width
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


VERIFIED_COLUMNS = [
    "Input Company Name",
    "Input Country",
    "Lead Type",
    "Xiaoman Company Name",
    "Domain",
    "Website Country",
    "Website",
    "B2B/B2C",
    "Verified Target",
    "Customer Type",
    "Is Competitor",
    "Goji Presence",
    "Rating",
    "P Priority",
    "Track Match",
    "Matched Track",
    "Evidence URL",
    "Primary Vertical",
    "Food/Supp Focus",
    "Rating Reason",
    "Outreach Angle",
    "Contact Count",
    "Email Count",
    "Contact Name",
    "Email",
    "Position",
]


def write_verified_xlsx(rows: list[dict], output_path: Path) -> None:
    """Write step4 verification output."""
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Verified"
    sheet.append(VERIFIED_COLUMNS)
    for row in rows:
        sheet.append([row.get(col, "") for col in VERIFIED_COLUMNS])
    sheet.freeze_panes = "A2"
    widths = [32, 16, 14, 32, 28, 18, 36, 12, 16, 18, 14, 14, 10, 12, 12, 20, 36, 18, 18, 40, 36, 12, 12, 32, 42, 32]
    for idx, width in enumerate(widths, start=1):
        sheet.column_dimensions[chr(ord("A") + idx - 1)].width = width
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


SALES_LEADS_COLUMNS = [
    "公司名",
    "国家",
    "网站",
    "公司介绍",
    "业务关联",
    "联系人",
    "职位",
    "邮箱",
    "电话",
    "有Contact Page",
    "邮件主题",
    "邮件正文",
    "WhatsApp/LinkedIn",
    "Follow-up 邮件",
    "状态",
]


def _estimated_row_height(row: dict) -> float:
    text_fields = ("公司介绍", "业务关联", "邮件正文", "WhatsApp/LinkedIn", "Follow-up 邮件")
    estimated_lines = 1
    for field in text_fields:
        text = str(row.get(field, "") or "")
        if not text:
            continue
        explicit_lines = text.count("\n") + 1
        wrapped_lines = (len(text) // 90) + 1
        estimated_lines = max(estimated_lines, explicit_lines, wrapped_lines)
    return min(180, max(36, estimated_lines * 15))


def write_sales_leads_xlsx(rows: list[dict], output_path: Path) -> None:
    """Write the final sales-facing lead sheet produced by step5."""
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sales Leads"

    sheet.append(SALES_LEADS_COLUMNS)
    header_fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
    header_alignment = Alignment(wrap_text=True, vertical="center")
    cell_alignment = Alignment(wrap_text=True, vertical="top")

    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = header_alignment

    for row in rows:
        sheet.append([row.get(col, "") for col in SALES_LEADS_COLUMNS])
        row_number = sheet.max_row
        sheet.row_dimensions[row_number].height = _estimated_row_height(row)

    sheet.freeze_panes = "A2"
    widths = [30, 16, 36, 44, 48, 24, 24, 38, 28, 16, 34, 64, 48, 56, 16]
    for idx, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(idx)].width = width

    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = cell_alignment if cell.row > 1 else header_alignment

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)

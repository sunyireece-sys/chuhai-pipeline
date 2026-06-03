"""
Step4: website verification + goji ingredient analysis (merged).

Reads 03_xiaoman.xlsx top-1 rows → fetches company websites → calls LLM judge
→ writes 04_verified.xlsx.
"""
from __future__ import annotations

import html
import logging
import re
import time
from pathlib import Path
from urllib.parse import urljoin, urlparse

from openpyxl import load_workbook

from llm_judge import JudgeInput, judge
from schema import write_verified_xlsx

log = logging.getLogger(__name__)

PAGE_PATHS = [
    "",
    "/about",
    "/about-us",
    "/products",
    "/catalog",
    "/shop",
    "/store",
    "/category",
    "/en/",
    "/ru/",
]
MAX_PAGE_CHARS = 5000
MAX_TOTAL_CHARS = 30000
MAX_URLS_PER_SITE = 10
GOJI_KEYWORDS = [
    "goji",
    "wolfberry",
    "lycium",
    "枸杞",
    "годжи",
    "годж",
    "kurt üzümü",
    "kurt uzumu",
    "kustovnice",
]
GOJI_KEYWORD_RE = re.compile("|".join(re.escape(k) for k in GOJI_KEYWORDS), re.IGNORECASE)
USER_AGENT = (
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
    "AppleWebKit/537.36 (KHTML, like Gecko) "
    "Chrome/146.0.0.0 Safari/537.36"
)


def _cell_text(value: object) -> str:
    return str(value or "").strip()


def _as_int(value: object) -> int:
    try:
        return int(value or 0)
    except Exception:
        return 0


def _read_xlsx_rows(path: Path) -> list[dict]:
    wb = load_workbook(path, read_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    header_row = next(rows, None)
    if not header_row:
        return []
    header = [_cell_text(h) for h in header_row]
    out = []
    for raw in rows:
        row = {}
        for idx, key in enumerate(header):
            row[key] = raw[idx] if idx < len(raw) and raw[idx] is not None else ""
        if any(_cell_text(value) for value in row.values()):
            out.append(row)
    return out


def _buyer_key(row: dict) -> tuple[str, str, str]:
    return (
        _cell_text(row.get("Input Company Name")),
        _cell_text(row.get("Input Country")),
        _cell_text(row.get("Input Lead Type")),
    )


def _unique_join(values: list[object]) -> str:
    seen: set[str] = set()
    out: list[str] = []
    for value in values:
        for part in _cell_text(value).split(";"):
            cleaned = part.strip()
            if cleaned and cleaned not in seen:
                seen.add(cleaned)
                out.append(cleaned)
    return "; ".join(out)


def _top1_company_rows(xiaoman_xlsx: Path) -> list[dict]:
    groups: dict[tuple[str, str, str], list[dict]] = {}
    for row in _read_xlsx_rows(xiaoman_xlsx):
        if _as_int(row.get("Match Rank")) != 1:
            continue
        groups.setdefault(_buyer_key(row), []).append(row)

    top_rows = []
    for rows in groups.values():
        first = dict(rows[0])
        first["Contact Name"] = _unique_join([row.get("Contact Name") for row in rows])
        first["Email"] = _unique_join([row.get("Email") for row in rows])
        first["Position"] = _unique_join([row.get("Position") for row in rows])
        top_rows.append(first)
    return top_rows


def _base_url(domain: str) -> str:
    value = _cell_text(domain)
    if not value:
        return ""
    if value.startswith(("http://", "https://")):
        parsed = urlparse(value)
        return f"{parsed.scheme}://{parsed.netloc}"
    return f"https://{value.strip('/')}"


def _html_to_text(content: str) -> str:
    content = re.sub(r"(?is)<(script|style|noscript).*?</\1>", " ", content)
    content = re.sub(r"(?s)<[^>]+>", " ", content)
    content = html.unescape(content)
    return re.sub(r"\s+", " ", content).strip()


def _response_text(resp: object) -> str:
    encoding = getattr(resp, "encoding", None)
    if encoding is None or str(encoding).lower() == "iso-8859-1":
        content = getattr(resp, "content", b"")
        try:
            return content.decode("utf-8")
        except UnicodeDecodeError:
            return content.decode("cp1251", errors="ignore")
    return getattr(resp, "text", "")


def _goji_contexts(text: str, *, radius: int = 200, limit: int = 5) -> list[str]:
    contexts: list[str] = []
    for match in GOJI_KEYWORD_RE.finditer(text or ""):
        start = max(0, match.start() - radius)
        end = min(len(text), match.end() + radius)
        contexts.append(text[start:end].strip())
        if len(contexts) >= limit:
            break
    return contexts


def _try_sitemap(client: object, base_url: str) -> list[str]:
    """Pull up to 20 product-like URLs from sitemap.xml."""
    try:
        resp = client.get(urljoin(base_url.rstrip("/") + "/", "sitemap.xml"))
        if resp.status_code != 200:
            return []
        urls = re.findall(r"<loc>([^<]+)</loc>", _response_text(resp))
    except Exception:
        return []
    product_like = [
        url
        for url in urls
        if any(
            key in url.lower()
            for key in ("product", "catalog", "shop", "goji", "berry", "годжи", "枸杞")
        )
    ]
    return product_like[:20]


def _fetch_website_pages(domain: str) -> tuple[str, list[str]]:
    import httpx

    base = _base_url(domain)
    if not base:
        return "", []

    chunks: list[str] = []
    fetched_urls: list[str] = []
    seen_urls: set[str] = set()

    def add_url(client: httpx.Client, url: str) -> None:
        if len(fetched_urls) >= MAX_URLS_PER_SITE or url in seen_urls:
            return
        seen_urls.add(url)
        try:
            resp = client.get(url)
            resp.raise_for_status()
        except Exception as exc:
            log.info("website fetch failed %s: %s", url, exc)
            return
        text = _html_to_text(_response_text(resp))
        if not text:
            return
        fetched_urls.append(str(resp.url))
        chunks.append(f"SOURCE_URL: {resp.url}\n{text[:MAX_PAGE_CHARS]}")

    with httpx.Client(
        follow_redirects=True,
        timeout=10.0,
        headers={"User-Agent": USER_AGENT},
    ) as client:
        for path in PAGE_PATHS:
            url = urljoin(base + "/", path.lstrip("/"))
            add_url(client, url)

        website_text = "\n\n".join(chunks)
        if not _goji_contexts(website_text):
            for url in _try_sitemap(client, base)[:5]:
                add_url(client, url)

    website_text = "\n\n".join(chunks)
    contexts = _goji_contexts(website_text)
    if contexts:
        parsed = urlparse(base)
        log.info("goji keyword hit on %s", parsed.netloc or base)
        context_block = "\n".join(f"GOJI_KEYWORD_CONTEXT: {ctx}" for ctx in contexts)
        website_text = f"{context_block}\n\n{website_text}"
    return website_text[:MAX_TOTAL_CHARS], fetched_urls


def fetch_website_text(domain: str) -> str:
    """Fetch selected website pages and return concatenated text."""
    text, _ = _fetch_website_pages(domain)
    return text


def run(xiaoman_xlsx: Path, output_path: Path, *, sleep_s: float = 2.0) -> Path:
    """Step4 main entry."""
    if output_path.exists():
        log.info("step4: %s already exists, skipping", output_path.name)
        return output_path
    if not xiaoman_xlsx.is_file():
        raise FileNotFoundError(f"step4: missing {xiaoman_xlsx}")

    top_rows = _top1_company_rows(xiaoman_xlsx)
    log.info("step4: %d top-1 companies to verify", len(top_rows))

    rows_out: list[dict] = []
    for idx, row in enumerate(top_rows, start=1):
        company_name = _cell_text(row.get("Xiaoman Company Name")) or _cell_text(row.get("Input Company Name"))
        website = _cell_text(row.get("Website"))
        domain = website or _cell_text(row.get("Domain"))
        website_text, fetched_urls = _fetch_website_pages(domain)
        evidence_fallback = fetched_urls[0] if fetched_urls else domain
        log.info("[%d/%d] step4 verify %r (%d chars)", idx, len(top_rows), company_name, len(website_text))

        verdict = judge(
            JudgeInput(
                company_name=company_name,
                country=_cell_text(row.get("Xiaoman Country")) or _cell_text(row.get("Input Country")),
                domain=_cell_text(row.get("Domain")) or domain,
                lead_type=_cell_text(row.get("Input Lead Type")),
                website_text=website_text,
            )
        )

        rows_out.append(
            {
                "Input Company Name": _cell_text(row.get("Input Company Name")),
                "Input Country": _cell_text(row.get("Input Country")),
                "Lead Type": _cell_text(row.get("Input Lead Type")),
                "Xiaoman Company Name": company_name,
                "Domain": _cell_text(row.get("Domain")),
                "Website Country": verdict.website_country,
                "Website": website,
                "B2B/B2C": verdict.b2b_or_b2c,
                "Verified Target": verdict.is_target,
                "Customer Type": verdict.customer_type,
                "Is Competitor": verdict.is_competitor,
                "Goji Presence": verdict.goji_presence,
                "Rating": verdict.rating,
                "P Priority": verdict.p_priority,
                "Track Match": verdict.track_match,
                "Matched Track": verdict.matched_track,
                "Evidence URL": verdict.evidence_url or evidence_fallback,
                "Primary Vertical": verdict.primary_vertical,
                "Food/Supp Focus": verdict.food_supplement_focus,
                "Rating Reason": verdict.rating_reason,
                "Outreach Angle": verdict.outreach_angle,
                "Contact Count": row.get("Contact Count", ""),
                "Email Count": row.get("Email Count", ""),
                "Contact Name": _cell_text(row.get("Contact Name")),
                "Email": _cell_text(row.get("Email")),
                "Position": _cell_text(row.get("Position")),
            }
        )

        if sleep_s > 0 and idx < len(top_rows):
            time.sleep(sleep_s)

    write_verified_xlsx(rows_out, output_path)
    log.info("step4: wrote %d verified rows to %s", len(rows_out), output_path.name)
    return output_path

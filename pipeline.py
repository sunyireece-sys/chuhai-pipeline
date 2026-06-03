"""
Haichu pipeline: keyword markdown → buyers.xlsx → Xiaoman → website verification.

Usage:
    python pipeline.py runs/xxx/01_keywords.md              # 全流程：step2→3→4→5→deploy
    python pipeline.py runs/xxx/01_keywords.md --skip-step5 # 跳过 step5
    python pipeline.py runs/xxx/01_keywords.md --no-deploy  # 不自动部署
    python pipeline.py runs/2026-04-08_goji/01_keywords.md
    python pipeline.py runs/2026-04-08_goji/01_keywords.md --skip-step3
    python pipeline.py runs/2026-04-08_goji/01_keywords.md --skip-contacts
    python pipeline.py runs/2026-04-08_goji/01_keywords.md --skip-step4
    python pipeline.py runs/2026-04-08_goji/01_keywords.md --max-queries 30

Step3 drives a real browser via Playwright (see xiaoman_playwright.py) so
the frontend computes signatures for us. First run opens a login page in
~/.xiaoman_playwright_profile — after that, the profile persists.
"""
from __future__ import annotations

import argparse
import datetime as dt
import json
import logging
import os
import random
import re
import shutil
import subprocess
import sys
import time
from collections import Counter, defaultdict
from pathlib import Path
from urllib.parse import urlparse

from dotenv import load_dotenv
from openpyxl import load_workbook

from buyer_extract import (
    CompanyCandidate,
    ExtractionContext,
    extract_from_serper,
    merge_by_domain,
    to_xlsx_rows,
)
from keyword_parser import KeywordPool, parse_markdown
from schema import write_buyers_xlsx, write_xiaoman_xlsx
from serper_search import SerperClient, SerperResult, normalize_country
from website_verify import run as run_website_verify
from xiaoman_playwright import XiaomanPlaywrightClient, XiaomanRateLimitError, annotate_and_rank_companies


# ----------------------------------------------------------------------
# CLI
# ----------------------------------------------------------------------


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Run the step2+step3+step4 pipeline.")
    parser.add_argument(
        "keywords",
        type=Path,
        nargs="?",
        help="Path to keyword markdown file (also defines the run dir as its parent).",
    )
    parser.add_argument(
        "--summary-only",
        type=Path,
        default=None,
        help="Recompute step3 summary from an existing run dir, then exit.",
    )
    parser.add_argument(
        "--run-dir",
        type=Path,
        default=None,
        help="Override the run dir. Defaults to the parent of the keyword file.",
    )
    parser.add_argument(
        "--max-queries",
        type=int,
        default=200,
        help="Cap the number of Serper queries (controls cost). Default 200.",
    )
    parser.add_argument(
        "--skip-step3",
        action="store_true",
        help="Skip the Xiaoman scrape step. Useful for testing the buyer extractor.",
    )
    parser.add_argument(
        "--skip-contacts",
        action="store_true",
        help="Skip contact fetch inside step3 and only write company-match rows.",
    )
    parser.add_argument(
        "--skip-step4",
        action="store_true",
        help="Skip website verification + LLM judging. Useful when OPENAI_API_KEY is not set.",
    )
    parser.add_argument(
        "--skip-step5",
        action="store_true",
        help="Skip profile enrichment (step5).",
    )
    parser.add_argument(
        "--no-deploy",
        action="store_true",
        help="Skip auto-deploy after step5.",
    )
    parser.add_argument(
        "--step5-limit",
        type=int,
        default=None,
        help="Pass --limit N to profile_enrich.py.",
    )
    parser.add_argument(
        "--results-per-query",
        type=int,
        default=10,
        help="How many organic results to fetch per query. Default 10.",
    )
    parser.add_argument(
        "--xiaoman-max-pages",
        type=int,
        default=1,
        help="Xiaoman result pages to fetch per buyer. 1 page = 20 matches. Default 1.",
    )
    parser.add_argument(
        "--xiaoman-sleep",
        type=float,
        default=None,
        help="Deprecated alias for --xiaoman-search-interval.",
    )
    parser.add_argument(
        "--xiaoman-search-interval",
        type=float,
        default=None,
        help="Target seconds between Xiaoman buyer searches. Default 8.0.",
    )
    parser.add_argument(
        "--xiaoman-jitter",
        type=float,
        default=0.5,
        help="Random +/- seconds around --xiaoman-search-interval. Default 0.5.",
    )
    parser.add_argument(
        "--xiaoman-batch-size",
        type=int,
        default=20,
        help="Pause after this many Xiaoman buyer searches. Set 0 to disable. Default 20.",
    )
    parser.add_argument(
        "--xiaoman-batch-pause",
        type=float,
        default=30.0,
        help="Seconds to pause after each Xiaoman batch. Default 30.",
    )
    return parser.parse_args()


def setup_logging(run_dir: Path) -> None:
    log_file = run_dir / "pipeline.log"
    run_dir.mkdir(parents=True, exist_ok=True)
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s  %(levelname)-7s %(message)s",
        handlers=[
            logging.FileHandler(log_file, encoding="utf-8"),
            logging.StreamHandler(sys.stdout),
        ],
        force=True,
    )


# ----------------------------------------------------------------------
# Step 2: Serper search
# ----------------------------------------------------------------------


def detect_anchor_lang(anchor: str) -> str:
    """Auto-detect anchor language by script/token. Returns ru, tr, or en."""
    if re.search(r"[А-Яа-яЁё]", anchor):
        return "ru"
    if re.search(r"[ĞÜŞİÖÇğüşıöç]", anchor):
        return "tr"
    turkish_tokens = {"toptan", "satis", "satış", "ithalatci", "ithalatçı", "meyvesi"}
    anchor_lower = anchor.lower()
    if any(token in anchor_lower for token in turkish_tokens):
        return "tr"
    return "en"


COUNTRY_LANGS_BY_GL = {
    "ru": ["ru"],
    "tr": ["tr"],
}


def country_langs(country: str) -> list[str]:
    """Return permitted query languages for a Serper country code."""
    _, gl = normalize_country(country)
    return COUNTRY_LANGS_BY_GL.get(gl, ["en"])


def build_queries(pool: KeywordPool, max_queries: int) -> list[tuple[str, str, str, str]]:
    """
    Build (query_text, country, modifier_dim, hl) tuples by compatible language.

    Order: A → B → C → D, then country. This way `--max-queries` truncation keeps
    coverage of MODIFIER-A first (the most reliable signal) before tailing into D.

    Neither anchor nor modifier is quoted: quoting either collapses Google recall
    to ~0 on narrow phrases. Loose matching may pull off-topic results, which the
    buyer_extract domain blacklist and downstream step3 are expected to filter.
    """
    queries: list[tuple[str, str, str, str]] = []
    skipped = 0
    for dim, mod_list in pool.modifiers_by_dim().items():
        for modifier in mod_list:
            for anchor in pool.anchors:
                anchor_lang = detect_anchor_lang(anchor)
                query_text = f"{anchor} {modifier}"
                for country in pool.countries:
                    if anchor_lang not in country_langs(country):
                        skipped += 1
                        continue
                    queries.append((query_text, country, dim, anchor_lang))

    if skipped:
        logging.info("build_queries: skipped %d combos due to language mismatch", skipped)

    if len(queries) > max_queries:
        logging.info(
            "truncating %d → %d queries (use --max-queries to raise the cap)",
            len(queries),
            max_queries,
        )
        queries = queries[:max_queries]
    return queries


def run_step2(
    pool: KeywordPool,
    run_dir: Path,
    max_queries: int,
    results_per_query: int,
) -> Path:
    buyers_xlsx = run_dir / "02_buyers.xlsx"
    raw_json = run_dir / "02_serper_raw.json"

    if buyers_xlsx.exists():
        logging.info("step2: %s already exists, skipping Serper", buyers_xlsx.name)
        return buyers_xlsx

    api_key = os.environ.get("SERPER_API_KEY", "")
    client = SerperClient(api_key=api_key)
    queries = build_queries(pool, max_queries=max_queries)
    logging.info("step2: %d queries to run against Serper", len(queries))

    all_candidates: list[CompanyCandidate] = []
    raw_records: list[dict] = []

    for idx, (query, country, dim, hl) in enumerate(queries, start=1):
        logging.info("[%d/%d] %s  (%s, MODIFIER-%s, hl=%s)", idx, len(queries), query, country, dim, hl)
        try:
            result: SerperResult = client.search(
                query=query, country=country, num=results_per_query, hl=hl
            )
        except Exception as exc:
            logging.error("  serper failed: %s — skipping this query", exc)
            continue

        raw_records.append(
            {
                "query": query,
                "country": country,
                "modifier": dim,
                "hl": hl,
                "country_code": result.country_code,
                "organic": result.organic,
            }
        )
        ctx = ExtractionContext(
            query=query,
            country_display=result.country_display,
            modifier_dim=dim,
        )
        new = extract_from_serper(result.organic, ctx)
        logging.info("  → %d candidates from %d organic", len(new), len(result.organic))
        all_candidates.extend(new)

    raw_json.write_text(json.dumps(raw_records, indent=2, ensure_ascii=False), encoding="utf-8")
    logging.info("wrote raw Serper responses to %s", raw_json.name)

    merged = merge_by_domain(all_candidates)
    rows = to_xlsx_rows(merged)
    write_buyers_xlsx(rows, buyers_xlsx)
    logging.info(
        "step2: wrote %d unique buyers to %s (%d Direct, %d Potential)",
        len(rows),
        buyers_xlsx.name,
        sum(1 for r in rows if r["Lead Type"] == "Direct Buyer"),
        sum(1 for r in rows if r["Lead Type"] == "Potential Buyer"),
    )
    return buyers_xlsx


# ----------------------------------------------------------------------
# Step 3: Xiaoman scrape (Playwright-driven)
# ----------------------------------------------------------------------


def read_buyer_names(buyers_xlsx: Path) -> list[dict]:
    """Load buyers.xlsx into buyer dicts keyed by header names."""
    wb = load_workbook(buyers_xlsx, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    header = [str(h).strip() if h is not None else "" for h in rows[0]]
    out = []
    for raw in rows[1:]:
        d = {header[i]: (raw[i] if i < len(raw) else "") for i in range(len(header))}
        name = (d.get("Company Name") or "").strip()
        if not name:
            continue
        out.append(
            {
                "Company Name": name,
                "Country": (d.get("Country") or "").strip(),
                "Lead Type": (d.get("Lead Type") or "").strip(),
                "Domain": (d.get("Domain") or "").strip(),
            }
        )
    return out


def read_xiaoman_xlsx(xiaoman_xlsx: Path) -> list[dict]:
    """Load 03_xiaoman.xlsx into row dicts keyed by the sheet header."""
    wb = load_workbook(xiaoman_xlsx, read_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    header_row = next(rows, None)
    if not header_row:
        return []
    header = [str(h).strip() if h is not None else "" for h in header_row]

    out = []
    for raw in rows:
        row = {}
        for idx, key in enumerate(header):
            value = raw[idx] if idx < len(raw) else ""
            row[key] = value if value is not None else ""
        if any(value != "" for value in row.values()):
            out.append(row)
    return out


def _pct(numerator: int, denominator: int) -> str:
    return f"{(100 * numerator / denominator):.0f}%" if denominator else "0%"


def _truthy(value: object) -> bool:
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return value != 0
    if isinstance(value, str):
        return value.strip().lower() in {"1", "true", "yes", "y"}
    return bool(value)


COUNTRY_ALIASES_TO_ISO2 = {
    "ARGENTINA": "AR",
    "AUSTRALIA": "AU",
    "AUSTRIA": "AT",
    "BELGIUM": "BE",
    "BRAZIL": "BR",
    "CANADA": "CA",
    "CHINA": "CN",
    "FRANCE": "FR",
    "GABON": "GA",
    "GERMANY": "DE",
    "ISRAEL": "IL",
    "ITALY": "IT",
    "NETHERLANDS": "NL",
    "THE NETHERLANDS": "NL",
    "RUSSIA": "RU",
    "RUSSIAN FEDERATION": "RU",
    "TURKEY": "TR",
    "TÜRKIYE": "TR",
    "UNITED KINGDOM": "GB",
    "UK": "GB",
    "GREAT BRITAIN": "GB",
    "UNITED KINGDOM OF GREAT BRITAIN AND NORTHERN IRELAND": "GB",
    "UNITED STATES": "US",
    "USA": "US",
    "U.S.": "US",
    "U.S.A.": "US",
    "UNITED STATES OF AMERICA": "US",
}


def _normalize_country_to_iso2(value: object) -> str:
    text = str(value or "").strip()
    if not text:
        return ""
    code = text.upper()
    if len(code) == 2 and code.isalpha():
        return code
    return COUNTRY_ALIASES_TO_ISO2.get(code, "")


def _domain_from_url_or_host(value: object) -> str:
    text = str(value or "").strip()
    if not text:
        return ""
    parsed = urlparse(text if "://" in text else f"//{text}")
    host = (parsed.netloc or parsed.path.split("/", 1)[0]).strip().lower()
    if "@" in host:
        host = host.rsplit("@", 1)[1]
    if ":" in host:
        host = host.split(":", 1)[0]
    if host.startswith("www."):
        host = host[4:]
    return host


def _collect_known_profile_domains(current_run_dir: Path) -> set[str]:
    """Scan other non-test runs for domains that already reached step5 profiles."""
    known: set[str] = set()
    runs_root = current_run_dir.parent
    if not runs_root.exists():
        return known
    for run_dir in runs_root.iterdir():
        if not run_dir.is_dir() or run_dir == current_run_dir:
            continue
        if run_dir.name.startswith("test_"):
            continue
        profiles_dir = run_dir / "05_profiles" / "profiles"
        if not profiles_dir.exists():
            continue
        for profile_path in profiles_dir.glob("*.json"):
            try:
                data = json.loads(profile_path.read_text(encoding="utf-8"))
            except Exception:
                continue
            company = data.get("company") if isinstance(data, dict) else {}
            if not isinstance(company, dict):
                company = {}
            for raw_domain in (
                data.get("domain") if isinstance(data, dict) else "",
                data.get("website") if isinstance(data, dict) else "",
                company.get("domain"),
                company.get("website"),
            ):
                domain = _domain_from_url_or_host(raw_domain)
                if domain:
                    known.add(domain)
    return known


def _buyer_key(row: dict) -> tuple[str, str, str]:
    return (
        (row.get("Company Name") or row.get("Input Company Name") or "").strip(),
        (row.get("Country") or row.get("Input Country") or "").strip(),
        (row.get("Lead Type") or row.get("Input Lead Type") or "").strip(),
    )


def _dedupe_buyers_by_key(buyers: list[dict]) -> list[dict]:
    seen: set[tuple[str, str, str]] = set()
    out: list[dict] = []
    dropped = 0
    for buyer in buyers:
        key = _buyer_key(buyer)
        if key in seen:
            dropped += 1
            continue
        seen.add(key)
        out.append(buyer)
    if dropped:
        logging.info("step3: dropped %d duplicate buyer rows before search", dropped)
    return out


def _buyer_key_id(key: tuple[str, str, str]) -> str:
    return json.dumps(list(key), ensure_ascii=False)


def _is_completed_xiaoman_buyer(row: dict) -> bool:
    if not row.get("Input Company Name"):
        return False
    return bool((row.get("Xiaoman Company Name") or "").strip())


def _completed_xiaoman_buyer_keys(rows_out: list[dict]) -> set[tuple[str, str, str]]:
    return {_buyer_key(row) for row in rows_out if _is_completed_xiaoman_buyer(row)}


def _step3_progress_path(xiaoman_xlsx: Path) -> Path:
    return xiaoman_xlsx.with_name("03_xiaoman_progress.json")


def _load_step3_progress(path: Path) -> dict:
    if not path.is_file():
        return {"buyers": {}}
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        logging.warning("step3 resume: failed to read %s; ignoring progress file", path.name)
        return {"buyers": {}}
    if not isinstance(data, dict):
        return {"buyers": {}}
    buyers = data.get("buyers")
    if not isinstance(buyers, dict):
        data["buyers"] = {}
    return data


def _save_step3_progress(path: Path, progress: dict) -> None:
    progress["updated_at"] = dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    path.write_text(json.dumps(progress, ensure_ascii=False, indent=2, sort_keys=True), encoding="utf-8")


def _progress_completed_buyer_keys(progress: dict) -> set[tuple[str, str, str]]:
    completed_statuses = {"completed", "no_matches"}
    out: set[tuple[str, str, str]] = set()
    for key_id, record in (progress.get("buyers") or {}).items():
        if not isinstance(record, dict) or record.get("status") not in completed_statuses:
            continue
        try:
            raw = json.loads(key_id)
        except Exception:
            continue
        if isinstance(raw, list) and len(raw) == 3:
            out.add(tuple(str(part or "") for part in raw))
    return out


def _set_progress_record(
    progress: dict,
    buyer: dict,
    *,
    status: str,
    buyer_index: int,
    total_buyers: int,
    match_count: int | None = None,
    detail: str = "",
) -> None:
    key = _buyer_key(buyer)
    record = {
        "company_name": key[0],
        "country": key[1],
        "lead_type": key[2],
        "status": status,
        "buyer_index": buyer_index,
        "total_buyers": total_buyers,
        "updated_at": dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    }
    if match_count is not None:
        record["match_count"] = match_count
    if detail:
        record["detail"] = detail
    progress.setdefault("buyers", {})[_buyer_key_id(key)] = record


def _infer_zero_match_keys_from_log(run_dir: Path, buyers: list[dict]) -> set[tuple[str, str, str]]:
    """Recover old 0-match attempts from pipeline.log for runs created before progress.json."""
    log_path = run_dir / "pipeline.log"
    if not log_path.is_file():
        return set()
    keys_by_name: dict[str, list[tuple[str, str, str]]] = defaultdict(list)
    for buyer in buyers:
        keys_by_name[(buyer.get("Company Name") or "").strip()].append(_buyer_key(buyer))
    try:
        text = log_path.read_text(encoding="utf-8", errors="ignore")
    except Exception:
        return set()
    inferred: set[tuple[str, str, str]] = set()
    for match in re.finditer(r"\[\d+/\d+\]\s+'(.+?)'\s+→\s+0 matches", text):
        name = match.group(1)
        keys = keys_by_name.get(name, [])
        if len(keys) == 1:
            inferred.add(keys[0])
    return inferred


def _step3_completed_keys_for_summary(
    xiaoman_xlsx: Path,
    rows_out: list[dict],
    buyers: list[dict],
) -> set[tuple[str, str, str]]:
    progress = _load_step3_progress(_step3_progress_path(xiaoman_xlsx))
    return (
        _completed_xiaoman_buyer_keys(rows_out)
        | _progress_completed_buyer_keys(progress)
        | _infer_zero_match_keys_from_log(xiaoman_xlsx.parent, buyers)
    )


def _top1_rows_by_buyer(rows_out: list[dict]) -> dict[tuple[str, str, str], list[dict]]:
    grouped: dict[tuple[str, str, str], list[dict]] = {}
    for row in rows_out:
        if row.get("Match Rank") != 1:
            continue
        key = _buyer_key(row)
        grouped.setdefault(key, []).append(row)
    return grouped


def _unique_match_count(rows_out: list[dict]) -> int:
    return len(
        {
            (
                (row.get("Input Company Name") or "").strip(),
                (row.get("Input Country") or "").strip(),
                row.get("Match Rank"),
            )
            for row in rows_out
            if row.get("Input Company Name") and row.get("Match Rank") not in ("", None)
        }
    )


def _blank_contact_fields() -> dict[str, str]:
    return {
        "Contact Name": "",
        "Position": "",
        "Email": "",
        "Email Quality": "",
        "Phone": "",
        "LinkedIn": "",
        "Confidence": "",
    }


def _build_step3_summary_markdown(
    buyers: list[dict], rows_out: list[dict], xiaoman_xlsx: Path
) -> str:
    """Render a markdown summary with the same metrics emitted to the log."""
    total_buyers = len(buyers)
    matched_buyer_keys = {_buyer_key(row) for row in rows_out if row.get("Input Company Name")}
    buyers_with_match = len(matched_buyer_keys)
    completed_buyer_keys = _step3_completed_keys_for_summary(xiaoman_xlsx, rows_out, buyers)
    completed_buyers = len(completed_buyer_keys)
    pending_buyers = max(0, total_buyers - completed_buyers)
    top_rows_by_buyer = _top1_rows_by_buyer(rows_out)
    top_rows = [rows[0] for rows in top_rows_by_buyer.values()]
    buyers_with_top1 = len(top_rows_by_buyer)
    buyers_without_top1 = [
        buyer for buyer in buyers
        if _buyer_key(buyer) in matched_buyer_keys and _buyer_key(buyer) not in top_rows_by_buyer
    ]
    top1_country_conflicts = sum(1 for r in top_rows if _truthy(r.get("Country Conflict")))
    top_with_website = sum(1 for r in top_rows if r["Domain"] or r["Website"])
    top_with_contacts = sum(1 for r in top_rows if (r["Contact Count"] or 0) > 0)
    top_with_contact_email = sum(
        1 for rows in top_rows_by_buyer.values() if any((row.get("Email") or "").strip() for row in rows)
    )
    total_contact_signals = sum(1 for r in rows_out if (r["Contact Count"] or 0) > 0)
    unique_match_count = _unique_match_count(rows_out)
    avg_matches_per_buyer = (unique_match_count / buyers_with_match) if buyers_with_match else 0.0

    top1_country_counts = Counter(
        (row.get("Xiaoman Country Code") or "").strip() or "(blank)" for row in top_rows
    )
    top1_country_top5 = top1_country_counts.most_common(5)

    buyers_with_comparable_country = 0
    buyers_country_mismatch = 0
    buyers_country_unmapped = 0
    for buyer in buyers:
        top_row = top_rows_by_buyer.get(_buyer_key(buyer), [{}])[0]
        xiaoman_country_code = _normalize_country_to_iso2(top_row.get("Xiaoman Country Code"))
        if not xiaoman_country_code:
            continue
        buyer_country = _normalize_country_to_iso2(buyer.get("Country"))
        if not buyer_country:
            buyers_country_unmapped += 1
            continue
        buyers_with_comparable_country += 1
        if buyer_country != xiaoman_country_code:
            buyers_country_mismatch += 1

    lines = [
        f"# Step3 Summary: {xiaoman_xlsx.name}",
        "",
        f"- Generated at: {dt.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
        f"- Source files: `{xiaoman_xlsx.name}`, `02_buyers.xlsx`",
        "",
        "| 指标 | 数值 | 说明 |",
        "| --- | --- | --- |",
        f"| 总行数 | {len(rows_out)} | `03_xiaoman.xlsx` 数据行数 |",
        f"| 已完成 buyer | {completed_buyers}/{total_buyers} ({_pct(completed_buyers, total_buyers)}) | 已写入至少 1 条 Xiaoman 公司候选的 buyer 数；用于断点续跑 |",
        f"| 未完成 buyer | {pending_buyers}/{total_buyers} ({_pct(pending_buyers, total_buyers)}) | 下次重跑同一 run 时会自动从这些 buyer 继续 |",
        f"| buyer 至少有候选 | {buyers_with_match}/{total_buyers} ({_pct(buyers_with_match, total_buyers)}) | 小满至少返回 1 家候选的 buyer 数 |",
        f"| eligible top-1 覆盖 | {buyers_with_top1}/{total_buyers} ({_pct(buyers_with_top1, total_buyers)}) | 通过名称/域名/国家质量门槛、可进入 step4 的 buyer 数 |",
        f"| 无 eligible top-1 | {len(buyers_without_top1)}/{total_buyers} ({_pct(len(buyers_without_top1), total_buyers)}) | 有候选但没有合格 top-1 的 buyer 数 |",
        f"| top-1 国家冲突 | {top1_country_conflicts}/{len(top_rows)} ({_pct(top1_country_conflicts, len(top_rows))}) | 强名称或中等域名+名称证据允许 rank 1，但 Xiaoman 国家与搜索国家不一致 |",
        f"| top-1 有网站 | {top_with_website}/{len(top_rows)} ({_pct(top_with_website, len(top_rows))}) | `Domain` 或 `Website` 非空 |",
        f"| top-1 有联系人计数 | {top_with_contacts}/{len(top_rows)} ({_pct(top_with_contacts, len(top_rows))}) | `Contact Count > 0` |",
        f"| top-1 有联系人邮箱 | {top_with_contact_email}/{total_buyers} ({_pct(top_with_contact_email, total_buyers)}) | top-1 行里 `Email` 非空的 buyer 数 |",
        f"| 全部行有联系人计数 | {total_contact_signals}/{len(rows_out)} ({_pct(total_contact_signals, len(rows_out))}) | 全部命中行里 `Contact Count > 0` |",
        f"| 平均每 buyer 命中数 | {avg_matches_per_buyer:.1f} | `唯一 (buyer, country, rank) 数 / buyers_with_match` |",
        f"| top-1 国家不匹配率 | {buyers_country_mismatch}/{buyers_with_comparable_country} ({_pct(buyers_country_mismatch, buyers_with_comparable_country)}) | buyer `Country` 归一到 ISO2 后，与 Xiaoman top-1 `Country Code` 不一致的比例 |",
        f"| top-1 国家不可比较 | {buyers_country_unmapped} | buyer `Country` 为空或无法归一到 ISO2，已从上面分母剔除 |",
        "",
        "## 国家分布 Top-5",
        "",
        "| Rank | Xiaoman Country Code | Count |",
        "| --- | --- | --- |",
    ]

    if top1_country_top5:
        for idx, (country_code, count) in enumerate(top1_country_top5, start=1):
            lines.append(f"| {idx} | {country_code} | {count} |")
    else:
        lines.append("| 1 | (no top-1 rows) | 0 |")

    if buyers_without_top1:
        lines.extend(
            [
                "",
                "## 无 Eligible Top-1 Buyer",
                "",
                "| Buyer | Country | Lead Type |",
                "| --- | --- | --- |",
            ]
        )
        for buyer in buyers_without_top1:
            lines.append(
                f"| {buyer.get('Company Name', '')} | {buyer.get('Country', '')} | {buyer.get('Lead Type', '')} |"
            )

    lines.extend(
        [
            "",
            "## Notes",
            "",
            "- top-1 国家不匹配率反映小满 top-1 匹配质量，不一致行需要人工/二次匹配复核。",
            "- `ZZ` 是 ISO 3166 的未知/未分配占位码；这里表示 Xiaoman top-1 国家码缺少可识别国家。",
        ]
    )
    return "\n".join(lines) + "\n"


def _log_step3_summary(buyers: list[dict], rows_out: list[dict], xiaoman_xlsx: Path) -> Path:
    """Print a quick data-quality readout and persist it as markdown."""
    summary_md = xiaoman_xlsx.with_name("03_xiaoman_summary.md")
    markdown = _build_step3_summary_markdown(buyers, rows_out, xiaoman_xlsx)
    summary_md.write_text(markdown, encoding="utf-8")

    total_buyers = len(buyers)
    buyers_with_match = len({_buyer_key(row) for row in rows_out if row.get("Input Company Name")})
    completed_buyer_keys = _step3_completed_keys_for_summary(xiaoman_xlsx, rows_out, buyers)
    completed_buyers = len(completed_buyer_keys)
    pending_buyers = max(0, total_buyers - completed_buyers)
    matched_buyer_keys = {_buyer_key(row) for row in rows_out if row.get("Input Company Name")}
    top_rows_by_buyer = _top1_rows_by_buyer(rows_out)
    top_rows = [rows[0] for rows in top_rows_by_buyer.values()]
    buyers_with_top1 = len(top_rows_by_buyer)
    buyers_without_top1 = [
        buyer for buyer in buyers
        if _buyer_key(buyer) in matched_buyer_keys and _buyer_key(buyer) not in top_rows_by_buyer
    ]
    top1_country_conflicts = sum(1 for r in top_rows if _truthy(r.get("Country Conflict")))
    top_with_website = sum(1 for r in top_rows if r["Domain"] or r["Website"])
    top_with_contacts = sum(1 for r in top_rows if (r["Contact Count"] or 0) > 0)
    top_with_contact_email = sum(
        1 for rows in top_rows_by_buyer.values() if any((row.get("Email") or "").strip() for row in rows)
    )
    total_contact_signals = sum(1 for r in rows_out if (r["Contact Count"] or 0) > 0)
    unique_match_count = _unique_match_count(rows_out)
    avg_matches_per_buyer = (unique_match_count / buyers_with_match) if buyers_with_match else 0.0
    top1_country_top5 = Counter(
        (row.get("Xiaoman Country Code") or "").strip() or "(blank)" for row in top_rows
    ).most_common(5)
    buyers_with_comparable_country = 0
    buyers_country_mismatch = 0
    buyers_country_unmapped = 0
    for buyer in buyers:
        xiaoman_country_code = _normalize_country_to_iso2(
            top_rows_by_buyer.get(_buyer_key(buyer), [{}])[0].get("Xiaoman Country Code")
        )
        if not xiaoman_country_code:
            continue
        buyer_country = _normalize_country_to_iso2(buyer.get("Country"))
        if not buyer_country:
            buyers_country_unmapped += 1
            continue
        buyers_with_comparable_country += 1
        if buyer_country != xiaoman_country_code:
            buyers_country_mismatch += 1

    logging.info("step3 summary ── %s", xiaoman_xlsx.name)
    logging.info("  总行数:             %d", len(rows_out))
    logging.info("  已完成 buyer:       %d/%d (%s)",
                 completed_buyers, total_buyers, _pct(completed_buyers, total_buyers))
    logging.info("  未完成 buyer:       %d/%d (%s)",
                 pending_buyers, total_buyers, _pct(pending_buyers, total_buyers))
    logging.info("  buyer 至少有候选:   %d/%d (%s)",
                 buyers_with_match, total_buyers, _pct(buyers_with_match, total_buyers))
    logging.info("  eligible top-1 覆盖:%d/%d (%s)",
                 buyers_with_top1, total_buyers, _pct(buyers_with_top1, total_buyers))
    logging.info("  无 eligible top-1:  %d/%d (%s)",
                 len(buyers_without_top1), total_buyers, _pct(len(buyers_without_top1), total_buyers))
    logging.info("  top-1 国家冲突:     %d/%d (%s)",
                 top1_country_conflicts, len(top_rows), _pct(top1_country_conflicts, len(top_rows)))
    logging.info("  top-1 有网站:       %d/%d (%s)",
                 top_with_website, len(top_rows), _pct(top_with_website, len(top_rows)))
    logging.info("  top-1 有联系人计数: %d/%d (%s)",
                 top_with_contacts, len(top_rows), _pct(top_with_contacts, len(top_rows)))
    logging.info("  top-1 有联系人邮箱: %d/%d (%s)",
                 top_with_contact_email, total_buyers, _pct(top_with_contact_email, total_buyers))
    logging.info("  全部行有联系人计数: %d/%d (%s)",
                 total_contact_signals, len(rows_out), _pct(total_contact_signals, len(rows_out)))
    logging.info("  平均每 buyer 命中数: %.1f", avg_matches_per_buyer)
    logging.info("  top-1 国家不匹配率: %d/%d (%s)",
                 buyers_country_mismatch, buyers_with_comparable_country,
                 _pct(buyers_country_mismatch, buyers_with_comparable_country))
    logging.info("  top-1 国家不可比较: %d", buyers_country_unmapped)
    if top1_country_top5:
        logging.info("  国家分布 top-5:     %s",
                     ", ".join(f"{code}={count}" for code, count in top1_country_top5))
    else:
        logging.info("  国家分布 top-5:     (no top-1 rows)")
    logging.info("  summary markdown:   %s", summary_md.name)
    return summary_md


def run_summary_only(run_dir: Path) -> Path:
    buyers_xlsx = run_dir / "02_buyers.xlsx"
    xiaoman_xlsx = run_dir / "03_xiaoman.xlsx"
    if not buyers_xlsx.is_file():
        raise FileNotFoundError(f"summary-only: missing {buyers_xlsx}")
    if not xiaoman_xlsx.is_file():
        raise FileNotFoundError(f"summary-only: missing {xiaoman_xlsx}")

    buyers = read_buyer_names(buyers_xlsx)
    rows_out = read_xiaoman_xlsx(xiaoman_xlsx)
    return _log_step3_summary(buyers, rows_out, xiaoman_xlsx)


def _company_match_row(
    buyer: dict,
    match_rank: int,
    company,
    *,
    name_similarity: float,
    country_match: bool,
    match_quality: str,
    top1_eligible: bool,
    country_conflict: bool,
) -> dict:
    return {
        "Input Company Name": buyer["Company Name"],
        "Input Country": buyer["Country"],
        "Input Lead Type": buyer["Lead Type"],
        "Match Rank": match_rank,
        "Xiaoman Company Name": company.company_name,
        "Xiaoman Country": company.country_name or company.country_cn_name,
        "Xiaoman Country Code": company.country_code,
        "Domain": company.domain,
        "Website": company.website,
        "Description": company.description,
        "Contact Count": company.contact_count,
        "Email Count": company.email_count,
        "Company Hash ID": company.company_hash_id,
        "Name Similarity": name_similarity,
        "Country Match": country_match,
        "Match Quality": match_quality,
        "Top-1 Eligible": top1_eligible,
        "Country Conflict": country_conflict,
        **_blank_contact_fields(),
    }


def _contact_fields(contact) -> dict[str, object]:
    return {
        "Contact Name": contact.contact_name,
        "Position": contact.position,
        "Email": "; ".join(contact.emails),
        "Email Quality": json.dumps(contact.email_quality, ensure_ascii=False, sort_keys=True),
        "Phone": "; ".join(contact.phone_numbers),
        "LinkedIn": contact.linkedin,
        "Confidence": contact.confidence,
    }


def _is_xiaoman_rate_limit_error(exc: Exception) -> bool:
    if isinstance(exc, XiaomanRateLimitError):
        return True
    text = str(exc).lower()
    return (
        "http-over-limit-captcha" in text
        or "rate-limit captcha" in text
        or "captcha not dismissed" in text
        or "訪問頻繁" in text
        or "访问频繁" in text
    )


def _is_xiaoman_session_closed_error(exc: Exception) -> bool:
    text = str(exc).lower()
    return "target page, context or browser has been closed" in text


def _target_interval(base_s: float, jitter_s: float) -> float:
    if base_s <= 0:
        return 0.0
    if jitter_s <= 0:
        return base_s
    return max(0.0, base_s + random.uniform(-jitter_s, jitter_s))


def run_step3(
    buyers_xlsx: Path,
    run_dir: Path,
    max_pages: int,
    search_interval_s: float,
    jitter_s: float,
    batch_size: int,
    batch_pause_s: float,
    *,
    skip_contacts: bool,
) -> Path:
    xiaoman_xlsx = run_dir / "03_xiaoman.xlsx"
    progress_path = _step3_progress_path(xiaoman_xlsx)
    buyers = read_buyer_names(buyers_xlsx)
    buyers = _dedupe_buyers_by_key(buyers)
    if not buyers:
        raise RuntimeError(f"step3: no buyer rows in {buyers_xlsx}")

    progress = _load_step3_progress(progress_path)
    buyer_positions = {_buyer_key(buyer): idx for idx, buyer in enumerate(buyers, start=1)}
    rows_out: list[dict] = []
    if xiaoman_xlsx.exists():
        rows_out = read_xiaoman_xlsx(xiaoman_xlsx)
        row_completed_keys = _completed_xiaoman_buyer_keys(rows_out)
        inferred_no_match_keys = _infer_zero_match_keys_from_log(run_dir, buyers)
        for key in row_completed_keys:
            progress.setdefault("buyers", {}).setdefault(
                _buyer_key_id(key),
                {
                    "company_name": key[0],
                    "country": key[1],
                    "lead_type": key[2],
                    "status": "completed",
                    "buyer_index": buyer_positions.get(key, 0),
                    "total_buyers": len(buyers),
                    "detail": "recovered from existing xlsx rows",
                },
            )
        for key in inferred_no_match_keys:
            progress.setdefault("buyers", {}).setdefault(
                _buyer_key_id(key),
                {
                    "company_name": key[0],
                    "country": key[1],
                    "lead_type": key[2],
                    "status": "no_matches",
                    "buyer_index": buyer_positions.get(key, 0),
                    "total_buyers": len(buyers),
                    "match_count": 0,
                    "detail": "recovered from pipeline.log",
                },
            )
        _save_step3_progress(progress_path, progress)
        completed_keys = row_completed_keys | _progress_completed_buyer_keys(progress)
        buyers_to_run = [buyer for buyer in buyers if _buyer_key(buyer) not in completed_keys]
        logging.info(
            "step3 resume: loaded %d existing rows from %s; %d/%d buyers already completed; %d remaining",
            len(rows_out),
            xiaoman_xlsx.name,
            len(completed_keys),
            len(buyers),
            len(buyers_to_run),
        )
        if not buyers_to_run:
            logging.info("step3 resume: all buyers already completed; refreshing summary only")
            _log_step3_summary(buyers, rows_out, xiaoman_xlsx)
            return xiaoman_xlsx
    else:
        buyers_to_run = buyers

    known_domains = _collect_known_profile_domains(run_dir)
    if known_domains:
        before = len(buyers_to_run)
        buyers_to_run = [
            buyer for buyer in buyers_to_run
            if _domain_from_url_or_host(buyer.get("Domain")) not in known_domains
        ]
        cross_dropped = before - len(buyers_to_run)
        if cross_dropped:
            logging.info("step3: dropped %d buyers already profiled in prior runs", cross_dropped)

    logging.info("step3: %d buyer rows to look up on xiaoman", len(buyers_to_run))
    if not buyers_to_run:
        logging.info("step3: no buyer rows remain after resume/dedup filters; refreshing summary only")
        write_xiaoman_xlsx(rows_out, xiaoman_xlsx)
        _log_step3_summary(buyers, rows_out, xiaoman_xlsx)
        return xiaoman_xlsx

    logging.info(
        "step3 throttle: target %.1fs/buyer, jitter +/- %.1fs, batch %d, batch pause %.1fs",
        search_interval_s,
        jitter_s,
        batch_size,
        batch_pause_s,
    )

    abort_reason = ""
    with XiaomanPlaywrightClient(captcha_action="abort") as client:
        for session_idx, buyer in enumerate(buyers_to_run, start=1):
            name = buyer["Company Name"]
            buyer_pos = buyer_positions.get(_buyer_key(buyer), session_idx)
            started_at = time.monotonic()
            try:
                matches = list(client.search(name, max_pages=max_pages))
            except Exception as exc:
                if _is_xiaoman_rate_limit_error(exc):
                    abort_reason = "Xiaoman captcha/rate-limit"
                    _set_progress_record(
                        progress,
                        buyer,
                        status="interrupted",
                        buyer_index=buyer_pos,
                        total_buyers=len(buyers),
                        detail=str(exc),
                    )
                    _save_step3_progress(progress_path, progress)
                    logging.error(
                        "[%d/%d] %r hit Xiaoman rate-limit/captcha: %s — aborting step3",
                        buyer_pos,
                        len(buyers),
                        name,
                        exc,
                    )
                    break
                if _is_xiaoman_session_closed_error(exc):
                    abort_reason = "Xiaoman browser/session closed"
                    _set_progress_record(
                        progress,
                        buyer,
                        status="interrupted",
                        buyer_index=buyer_pos,
                        total_buyers=len(buyers),
                        detail=str(exc),
                    )
                    _save_step3_progress(progress_path, progress)
                    logging.error(
                        "[%d/%d] %r failed because Xiaoman browser/session closed: %s — aborting step3",
                        buyer_pos,
                        len(buyers),
                        name,
                        exc,
                    )
                    break
                _set_progress_record(
                    progress,
                    buyer,
                    status="error",
                    buyer_index=buyer_pos,
                    total_buyers=len(buyers),
                    detail=str(exc),
                )
                _save_step3_progress(progress_path, progress)
                logging.error("[%d/%d] %r failed: %s — skipping", buyer_pos, len(buyers), name, exc)
                continue
            logging.info("[%d/%d] %r → %d matches", buyer_pos, len(buyers), name, len(matches))
            if not matches:
                _set_progress_record(
                    progress,
                    buyer,
                    status="no_matches",
                    buyer_index=buyer_pos,
                    total_buyers=len(buyers),
                    match_count=0,
                )
                _save_step3_progress(progress_path, progress)

            ranked_matches = annotate_and_rank_companies(
                input_name=name,
                input_country=buyer["Country"],
                companies=matches,
            )
            contacts_by_rank: dict[int, list] = {}
            top_match = next((company for rank, company, _, _, _, _, _ in ranked_matches if rank == 1), None)
            if (
                top_match is not None
                and not skip_contacts
                and (top_match.contact_count or 0) > 0
            ):
                contact_keyword = (top_match.company_name or "").strip() or name
                try:
                    contacts = client.fetch_contacts(contact_keyword, max_contacts_pages=1)
                    if contacts:
                        contacts_by_rank[1] = contacts
                    logging.info(
                        "[%d/%d] %r top-1 contacts via %r → %d",
                        buyer_pos,
                        len(buyers),
                        name,
                        contact_keyword,
                        len(contacts),
                    )
                except Exception as exc:
                    logging.error(
                        "[%d/%d] %r top-1 contacts via %r failed: %s — keeping blank contact cells",
                        buyer_pos,
                        len(buyers),
                        name,
                        contact_keyword,
                        exc,
                    )

            if matches and top_match is None:
                logging.warning("[%d/%d] %r has no eligible top-1 after match quality filter", buyer_pos, len(buyers), name)

            rows_before_buyer = len(rows_out)
            for rank, company, similarity, country_match, match_quality, top1_eligible, country_conflict in ranked_matches:
                base_row = _company_match_row(
                    buyer,
                    rank,
                    company,
                    name_similarity=similarity,
                    country_match=country_match,
                    match_quality=match_quality,
                    top1_eligible=top1_eligible,
                    country_conflict=country_conflict,
                )
                if rank == 1 and contacts_by_rank.get(rank):
                    for contact in contacts_by_rank[rank]:
                        rows_out.append({**base_row, **_contact_fields(contact)})
                else:
                    rows_out.append(base_row)

            if len(rows_out) > rows_before_buyer:
                _set_progress_record(
                    progress,
                    buyer,
                    status="completed",
                    buyer_index=buyer_pos,
                    total_buyers=len(buyers),
                    match_count=len(matches),
                )
                _save_step3_progress(progress_path, progress)

            if session_idx < len(buyers_to_run):
                if batch_size > 0 and session_idx % batch_size == 0 and batch_pause_s > 0:
                    logging.info(
                        "xiaoman throttle: completed %d buyer searches; pausing %.1fs",
                        session_idx,
                        batch_pause_s,
                    )
                    time.sleep(batch_pause_s)
                else:
                    target_s = _target_interval(search_interval_s, jitter_s)
                    elapsed_s = time.monotonic() - started_at
                    sleep_s = max(0.0, target_s - elapsed_s)
                    if sleep_s > 0:
                        time.sleep(sleep_s)

    if abort_reason and not _completed_xiaoman_buyer_keys(rows_out):
        raise RuntimeError(f"step3 aborted by {abort_reason} before any rows were collected")
    if abort_reason:
        logging.warning("step3 wrote partial results because %s interrupted the run", abort_reason)
    write_xiaoman_xlsx(rows_out, xiaoman_xlsx)
    _log_step3_summary(buyers, rows_out, xiaoman_xlsx)
    return xiaoman_xlsx


def run_step4(xiaoman_xlsx: Path, run_dir: Path) -> Path:
    verified_xlsx = run_dir / "04_verified.xlsx"
    if verified_xlsx.exists():
        if xiaoman_xlsx.exists() and xiaoman_xlsx.stat().st_mtime > verified_xlsx.stat().st_mtime:
            logging.info("step4: 03_xiaoman.xlsx 比 04_verified.xlsx 新，删除旧输出重跑")
            verified_xlsx.unlink()
            profiles_dir = run_dir / "05_profiles"
            if profiles_dir.exists():
                shutil.rmtree(profiles_dir)
                logging.info("step4: 同时删除 05_profiles/ 以强制 step5 重跑")
        else:
            logging.info("step4: %s 已存在且为最新，跳过", verified_xlsx.name)
            return verified_xlsx
    return run_website_verify(xiaoman_xlsx, verified_xlsx)


def run_step5(verified_xlsx: Path, run_dir: Path, limit: int | None = None) -> Path:
    profiles_root = run_dir / "05_profiles"
    profiles_dir = run_dir / "05_profiles" / "profiles"
    existing_profiles = list(profiles_dir.glob("*.json")) if profiles_dir.is_dir() else []
    if existing_profiles:
        newest_profile = max(existing_profiles, key=lambda p: p.stat().st_mtime)
        if verified_xlsx.stat().st_mtime > newest_profile.stat().st_mtime:
            logging.info("step5: 04_verified.xlsx 比现有 profile 新，删除 05_profiles/ 重跑")
            shutil.rmtree(profiles_root)
            existing_profiles = []
    if existing_profiles:
        logging.info(
            "step5: found %d existing profile JSON files; profile_enrich.py will skip existing leads",
            len(existing_profiles),
        )

    script = Path(__file__).parent / "profile_enrich.py"
    cmd = [sys.executable, str(script), str(verified_xlsx)]
    if limit is not None:
        cmd.extend(["--limit", str(limit)])

    logging.info("step5: running %s", " ".join(cmd))
    result = subprocess.run(cmd, cwd=Path(__file__).parent)
    if result.returncode != 0:
        logging.error("step5 failed with exit code %d", result.returncode)
        raise RuntimeError(f"step5 failed with exit code {result.returncode}")
    return run_dir / "05_profiles"


def run_deploy(project_root: Path) -> None:
    deploy_sh = project_root / "deploy.sh"
    if not deploy_sh.is_file():
        logging.warning("deploy.sh not found, skipping deploy")
        return
    try:
        subprocess.run(["bash", str(deploy_sh)], cwd=project_root, check=True)
    except subprocess.CalledProcessError as exc:
        logging.error("deploy failed with exit code %d; local pipeline data is still available", exc.returncode)


# ----------------------------------------------------------------------
# Main
# ----------------------------------------------------------------------


def main() -> None:
    args = parse_args()

    if args.summary_only is not None:
        run_dir = args.summary_only.expanduser().resolve()
        setup_logging(run_dir)
        logging.info("summary-only run dir: %s", run_dir)
        summary_md = run_summary_only(run_dir)
        logging.info("summary-only complete")
        print(summary_md.read_text(encoding="utf-8"))
        return

    if args.keywords is None:
        print("keywords path is required unless --summary-only is used", file=sys.stderr)
        sys.exit(2)

    keywords_path = args.keywords.expanduser().resolve()
    if not keywords_path.is_file():
        print(f"keyword file not found: {keywords_path}", file=sys.stderr)
        sys.exit(2)

    run_dir = (args.run_dir or keywords_path.parent).expanduser().resolve()
    setup_logging(run_dir)

    # Load .env from the script's own directory (not cwd) so it works regardless
    # of where the user invokes the script from.
    load_dotenv(Path(__file__).parent / ".env")

    logging.info("keyword file: %s", keywords_path)
    logging.info("run dir:      %s", run_dir)

    pool = parse_markdown(keywords_path)
    logging.info(
        "parsed pool: %d countries, %d anchors, %d modifiers (A=%d B=%d C=%d D=%d)",
        len(pool.countries),
        len(pool.anchors),
        pool.total_modifiers(),
        len(pool.modifier_a),
        len(pool.modifier_b),
        len(pool.modifier_c),
        len(pool.modifier_d),
    )

    buyers_xlsx = run_step2(
        pool=pool,
        run_dir=run_dir,
        max_queries=args.max_queries,
        results_per_query=args.results_per_query,
    )

    if args.skip_step3:
        logging.info("step3 skipped (--skip-step3 set)")
        xiaoman_xlsx = run_dir / "03_xiaoman.xlsx"
    else:
        xiaoman_search_interval = (
            args.xiaoman_search_interval
            if args.xiaoman_search_interval is not None
            else args.xiaoman_sleep
            if args.xiaoman_sleep is not None
            else 8.0
        )
        xiaoman_xlsx = run_step3(
            buyers_xlsx=buyers_xlsx,
            run_dir=run_dir,
            max_pages=args.xiaoman_max_pages,
            search_interval_s=xiaoman_search_interval,
            jitter_s=args.xiaoman_jitter,
            batch_size=args.xiaoman_batch_size,
            batch_pause_s=args.xiaoman_batch_pause,
            skip_contacts=args.skip_contacts,
        )

    if args.skip_step4:
        logging.info("step4 skipped (--skip-step4 set)")
    elif not xiaoman_xlsx.is_file():
        logging.info("step4 skipped (no %s found)", xiaoman_xlsx.name)
    else:
        run_step4(xiaoman_xlsx, run_dir)

    verified_xlsx = run_dir / "04_verified.xlsx"
    if args.skip_step4 or not verified_xlsx.is_file():
        logging.info("step5 skipped (no 04_verified.xlsx)")
    elif args.skip_step5:
        logging.info("step5 skipped (--skip-step5 set)")
    else:
        run_step5(verified_xlsx, run_dir, limit=args.step5_limit)
        if not args.no_deploy:
            run_deploy(Path(__file__).parent)
        else:
            logging.info("deploy skipped (--no-deploy set)")
    logging.info("pipeline complete")


if __name__ == "__main__":
    main()

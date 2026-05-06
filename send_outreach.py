"""
Send Step5 cold outreach through SMTP.

Default mode is dry-run: messages are sent to a test recipient while preserving
the original customer recipient in the subject/body and send log.
"""
from __future__ import annotations

import argparse
import datetime as dt
import json
import os
import re
import smtplib
import time
from dataclasses import dataclass
from email.message import EmailMessage
from email.utils import formataddr
from pathlib import Path
from typing import Callable, Iterable
from urllib.parse import urlparse

from dotenv import load_dotenv

SALES_LEADS_XLSX = "05_sales_leads.xlsx"
SEND_LOG_JSONL = "06_send_log.jsonl"
SEND_STATUS_HEADER = "发送状态"
SENT_STATUS = "sent"
SENDER_DISPLAY_NAME = "Nicky (Bairuiyuan Goji)"
DEFAULT_SLEEP_SECONDS = 5.0
BLOCKED_EMAIL_LOCAL_PARTS = {
    "abuse",
    "postmaster",
    "privacyofficer",
}
OUTREACH_KEYS = (
    "cold_email_subject",
    "cold_email_body",
    "whatsapp_or_linkedin_message",
    "follow_up_email",
)


@dataclass
class SendConfig:
    smtp_host: str
    smtp_port: int
    smtp_user: str
    smtp_pass: str
    live: bool = False
    test_recipient: str = ""
    sleep_s: float = DEFAULT_SLEEP_SECONDS


@dataclass
class ProfileRecord:
    slug: str
    path: Path
    data: dict


class ConfigError(RuntimeError):
    pass


def _cell_text(value: object) -> str:
    return str(value or "").strip()


def _load_json(path: Path) -> dict | None:
    try:
        if path.is_file():
            data = json.loads(path.read_text(encoding="utf-8"))
            return data if isinstance(data, dict) else None
    except Exception:
        return None
    return None


def _normalize_key(value: object) -> str:
    return re.sub(r"\s+", " ", _cell_text(value).lower()).strip()


def _base_url(value: object) -> str:
    text = _cell_text(value)
    if not text:
        return ""
    if not text.startswith(("http://", "https://")):
        text = "https://" + text.strip("/")
    parsed = urlparse(text)
    if not parsed.netloc:
        return ""
    return f"{parsed.scheme}://{parsed.netloc}".lower().rstrip("/")


def _profile_lookup_keys(profile: dict) -> Iterable[tuple[str, str]]:
    company = profile.get("company") if isinstance(profile.get("company"), dict) else {}
    for value in (
        company.get("display_name"),
        company.get("input_company_name"),
    ):
        key = _normalize_key(value)
        if key:
            yield ("name", key)
    for value in (company.get("website"),):
        key = _base_url(value)
        if key:
            yield ("site", key)


def _row_lookup_keys(row: dict) -> Iterable[tuple[str, str]]:
    for value in (row.get("公司名"),):
        key = _normalize_key(value)
        if key:
            yield ("name", key)
    for value in (row.get("网站"),):
        key = _base_url(value)
        if key:
            yield ("site", key)


def load_profiles(output_dir: Path) -> dict[tuple[str, str], ProfileRecord]:
    lookup: dict[tuple[str, str], ProfileRecord] = {}
    profiles_dir = output_dir / "profiles"
    if not profiles_dir.is_dir():
        return lookup
    for path in sorted(profiles_dir.glob("*.json")):
        data = _load_json(path)
        if not data:
            continue
        record = ProfileRecord(slug=path.stem, path=path, data=data)
        for key in _profile_lookup_keys(data):
            lookup.setdefault(key, record)
    return lookup


def _find_profile(row: dict, profiles_by_key: dict[tuple[str, str], ProfileRecord]) -> ProfileRecord | None:
    for key in _row_lookup_keys(row):
        record = profiles_by_key.get(key)
        if record is not None:
            return record
    return None


def _clean_email(value: object) -> str:
    return _cell_text(value).strip(" ,;<>")


def _is_blocked_email(value: str) -> bool:
    local_part = value.split("@", 1)[0].lower()
    return local_part in BLOCKED_EMAIL_LOCAL_PARTS


def _unique_business_emails(values: object) -> list[str]:
    emails = values if isinstance(values, list) else [values]
    out: list[str] = []
    seen: set[str] = set()
    if not isinstance(emails, list):
        emails = [emails]
    for email in emails:
        value = _clean_email(email)
        key = value.lower()
        if not value or "@" not in value or key in seen or _is_blocked_email(value):
            continue
        seen.add(key)
        out.append(value)
    return out


def _first_contact_email(output_dir: Path, slug: str) -> str:
    contacts = _load_json(output_dir / "contacts" / f"{slug}.json")
    if not isinstance(contacts, dict):
        return ""
    emails = _unique_business_emails(contacts.get("emails") or [])
    return emails[0] if emails else ""


def _profile_outreach(profile: dict) -> dict:
    outreach = profile.get("outreach") if isinstance(profile.get("outreach"), dict) else {}
    return outreach if isinstance(outreach, dict) else {}


def _redact_secret(text: object, secret: str) -> str:
    out = str(text or "")
    if secret:
        out = out.replace(secret, "***")
    return out


def _short_error(message: str, limit: int = 120) -> str:
    text = re.sub(r"\s+", " ", message).strip()
    return text[:limit]


def load_send_config(*, live: bool, test_recipient: str, sleep_s: float) -> SendConfig:
    smtp_host = _cell_text(os.environ.get("SMTP_HOST"))
    smtp_port_raw = _cell_text(os.environ.get("SMTP_PORT"))
    smtp_user = _cell_text(os.environ.get("SMTP_USER"))
    smtp_pass = _cell_text(os.environ.get("SMTP_PASS"))
    missing = [
        name
        for name, value in (
            ("SMTP_HOST", smtp_host),
            ("SMTP_PORT", smtp_port_raw),
            ("SMTP_USER", smtp_user),
            ("SMTP_PASS", smtp_pass),
        )
        if not value
    ]
    if missing:
        raise ConfigError(f"Missing SMTP environment variables: {', '.join(missing)}")
    try:
        smtp_port = int(smtp_port_raw)
    except ValueError as exc:
        raise ConfigError("SMTP_PORT must be an integer") from exc
    if not live and not _cell_text(test_recipient):
        raise ConfigError("--test-recipient is required in dry-run mode")
    if live and _cell_text(test_recipient):
        raise ConfigError("--test-recipient is only used in dry-run mode")
    return SendConfig(
        smtp_host=smtp_host,
        smtp_port=smtp_port,
        smtp_user=smtp_user,
        smtp_pass=smtp_pass,
        live=live,
        test_recipient=_cell_text(test_recipient),
        sleep_s=sleep_s,
    )


def _ensure_status_column(sheet) -> int:
    from openpyxl.utils import get_column_letter

    headers = [_cell_text(cell.value) for cell in sheet[1]]
    if SEND_STATUS_HEADER in headers:
        return headers.index(SEND_STATUS_HEADER) + 1
    column_idx = len(headers) + 1
    sheet.cell(row=1, column=column_idx, value=SEND_STATUS_HEADER)
    sheet.column_dimensions[get_column_letter(column_idx)].width = 18
    for row_idx in range(2, sheet.max_row + 1):
        sheet.cell(row=row_idx, column=column_idx, value="pending")
    return column_idx


def _headers(sheet) -> list[str]:
    return [_cell_text(cell.value) for cell in sheet[1]]


def _row_dict(sheet, row_idx: int, headers: list[str]) -> dict:
    return {
        header: sheet.cell(row=row_idx, column=col_idx).value
        for col_idx, header in enumerate(headers, start=1)
        if header
    }


def _log_record(log_path: Path, record: dict) -> None:
    log_path.parent.mkdir(parents=True, exist_ok=True)
    with log_path.open("a", encoding="utf-8") as handle:
        handle.write(json.dumps(record, ensure_ascii=False) + "\n")


def _send_message(
    *,
    config: SendConfig,
    original_to: str,
    actual_to: str,
    subject: str,
    body: str,
    smtp_factory: Callable[..., object] | None = None,
) -> str:
    message = EmailMessage()
    message["From"] = formataddr((SENDER_DISPLAY_NAME, config.smtp_user))
    message["To"] = actual_to
    message["Reply-To"] = config.smtp_user
    message["Subject"] = subject
    message.set_content(body)

    use_ssl = config.smtp_port == 465
    if smtp_factory is None:
        smtp_factory = smtplib.SMTP_SSL if use_ssl else smtplib.SMTP
    with smtp_factory(config.smtp_host, config.smtp_port) as smtp:
        if not use_ssl:
            smtp.starttls()
        smtp.login(config.smtp_user, config.smtp_pass)
        response = smtp.send_message(message)
    if response:
        return json.dumps(response, ensure_ascii=False, default=str)
    return "250 OK"


def _build_message(*, live: bool, test_recipient: str, original_to: str, subject: str, body: str) -> tuple[str, str, str]:
    if live:
        return original_to, subject, body
    dry_subject = f"[DRY-RUN to {original_to}] {subject}"
    dry_body = f"--- DRY-RUN, original recipient: {original_to} ---\n\n{body}"
    return test_recipient, dry_subject, dry_body


def _status_log_base(*, slug: str, mode: str, original_to: str, actual_to: str, subject: str) -> dict:
    return {
        "timestamp": dt.datetime.now().isoformat(timespec="seconds"),
        "slug": slug,
        "mode": mode,
        "original_to": original_to,
        "actual_to": actual_to,
        "subject": subject,
    }


def run_send(
    output_dir: Path,
    config: SendConfig,
    *,
    smtp_factory: Callable[..., object] = smtplib.SMTP,
    reset_status: bool = False,
) -> dict:
    from openpyxl import load_workbook

    sales_xlsx = output_dir / SALES_LEADS_XLSX
    if not sales_xlsx.is_file():
        raise FileNotFoundError(f"missing sales leads xlsx: {sales_xlsx}")

    workbook = load_workbook(sales_xlsx)
    sheet = workbook.active
    status_col = _ensure_status_column(sheet)
    if reset_status:
        for row_idx in range(2, sheet.max_row + 1):
            sheet.cell(row=row_idx, column=status_col, value="pending")
    headers = _headers(sheet)
    profiles_by_key = load_profiles(output_dir)
    log_path = output_dir / SEND_LOG_JSONL
    mode = "live" if config.live else "dry-run"

    summary = {"sent": 0, "skipped": 0, "errors": 0}
    for row_idx in range(2, sheet.max_row + 1):
        status_cell = sheet.cell(row=row_idx, column=status_col)
        current_status = _cell_text(status_cell.value)
        if current_status == SENT_STATUS:
            continue

        row = _row_dict(sheet, row_idx, headers)
        profile_record = _find_profile(row, profiles_by_key)
        profile = profile_record.data if profile_record else {}
        slug = profile_record.slug if profile_record else f"row-{row_idx}"
        if _cell_text(profile.get("outreach_mode")).lower() != "sales":
            status_cell.value = "pending" if not current_status else current_status
            continue

        original_to = _first_contact_email(output_dir, slug)
        outreach = _profile_outreach(profile)
        subject = _cell_text(outreach.get("cold_email_subject")) or _cell_text(row.get("邮件主题"))
        body = _cell_text(outreach.get("cold_email_body")) or _cell_text(row.get("邮件正文"))

        if not original_to:
            status_cell.value = "skip_no_email"
            summary["skipped"] += 1
            _log_record(
                log_path,
                {
                    **_status_log_base(slug=slug, mode=mode, original_to="", actual_to="", subject=subject),
                    "status": "skip_no_email",
                    "smtp_response": None,
                    "error": None,
                },
            )
            continue
        if not subject or not body:
            status_cell.value = "skip_no_content"
            summary["skipped"] += 1
            actual_to = original_to if config.live else config.test_recipient
            _log_record(
                log_path,
                {
                    **_status_log_base(
                        slug=slug,
                        mode=mode,
                        original_to=original_to,
                        actual_to=actual_to,
                        subject=subject,
                    ),
                    "status": "skip_no_content",
                    "smtp_response": None,
                    "error": None,
                },
            )
            continue

        actual_to, send_subject, send_body = _build_message(
            live=config.live,
            test_recipient=config.test_recipient,
            original_to=original_to,
            subject=subject,
            body=body,
        )
        log_base = _status_log_base(
            slug=slug,
            mode=mode,
            original_to=original_to,
            actual_to=actual_to,
            subject=send_subject,
        )
        try:
            smtp_response = _send_message(
                config=config,
                original_to=original_to,
                actual_to=actual_to,
                subject=send_subject,
                body=send_body,
                smtp_factory=smtp_factory,
            )
            status_cell.value = SENT_STATUS
            summary["sent"] += 1
            _log_record(
                log_path,
                {
                    **log_base,
                    "status": SENT_STATUS,
                    "smtp_response": _redact_secret(smtp_response, config.smtp_pass),
                    "error": None,
                },
            )
        except Exception as exc:
            error = _redact_secret(f"{type(exc).__name__}: {exc}", config.smtp_pass)
            status_cell.value = f"error: {_short_error(error)}"
            summary["errors"] += 1
            _log_record(
                log_path,
                {
                    **log_base,
                    "status": "error",
                    "smtp_response": None,
                    "error": error,
                },
            )

        if config.sleep_s > 0 and row_idx < sheet.max_row:
            time.sleep(config.sleep_s)

    workbook.save(sales_xlsx)
    return summary


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Send Step5 outreach emails through Outlook SMTP.")
    parser.add_argument("output_dir", type=Path, help="Path to runs/<date>/05_profiles/.")
    parser.add_argument("--test-recipient", default="", help="Dry-run recipient. Required unless --live is set.")
    parser.add_argument("--live", action="store_true", help="Actually send to customer recipients.")
    parser.add_argument(
        "--i-confirm-live-send",
        action="store_true",
        help="Required together with --live to prevent accidental customer sends.",
    )
    parser.add_argument("--reset", action="store_true", help="Reset the xlsx send status column to pending before sending.")
    parser.add_argument("--sleep", type=float, default=DEFAULT_SLEEP_SECONDS, help="Seconds to sleep after each send.")
    return parser.parse_args()


def main() -> None:
    load_dotenv()
    args = parse_args()
    if args.live and not args.i_confirm_live_send:
        raise SystemExit("--live requires --i-confirm-live-send")
    config = load_send_config(live=args.live, test_recipient=args.test_recipient, sleep_s=args.sleep)
    summary = run_send(args.output_dir, config, reset_status=args.reset)
    print(json.dumps(summary, ensure_ascii=False))


if __name__ == "__main__":
    main()
